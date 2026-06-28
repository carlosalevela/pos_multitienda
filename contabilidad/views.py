from decimal import Decimal

from django.db.models import Count, DecimalField, F, Sum
from django.db.models.functions import Coalesce, TruncDate, TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import generics
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from caja.models import MovimientoCaja, SesionCaja
from core.permissions import (
    EsAdminOSupervisor, EsAdminSupervisorOCajero,
    es_superadmin, get_empresa, scope_qs,
)
from devoluciones.models import Devolucion
from productos.models import MovimientoInventario
from ventas.models import DetalleVenta, Venta

from .models import Gasto
from .serializers import GastoSerializer


CATEGORIAS_SOLO_ADMIN = {
    'arriendo', 'nomina', 'servicios', 'mercancia',
    'recibo', 'proveedor', 'impuesto', 'administrativo',
}

MESES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


def _cajas_abiertas(request, tienda_id, fecha_fin):
    """
    Devuelve lista de SesionCaja abiertas si fecha_fin >= hoy.
    Vacía para reportes históricos (datos ya cerrados, sin riesgo).
    Usado para bloquear exports y advertir en reportes JSON.
    """
    from datetime import date as _date
    today = timezone.now().date()
    try:
        fd = _date.fromisoformat(str(fecha_fin)) if fecha_fin else today
    except (ValueError, TypeError):
        return []
    if fd < today:
        return []

    qs = SesionCaja.objects.filter(estado="abierta").select_related("tienda", "empleado")
    if not es_superadmin(request):
        qs = qs.filter(tienda__empresa=get_empresa(request))
    if tienda_id:
        qs = qs.filter(tienda_id=tienda_id)

    return [
        {
            "sesion_id": s.id,
            "tienda":    s.tienda.nombre,
            "empleado":  f"{s.empleado.nombre} {s.empleado.apellido}" if s.empleado else "—",
            "desde":     s.fecha_apertura.strftime("%H:%M"),
        }
        for s in qs
    ]


# ── Gastos ────────────────────────────────────────────────────
class GastoListCreateView(generics.ListCreateAPIView):
    serializer_class = GastoSerializer

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAuthenticated()]
        return [EsAdminSupervisorOCajero()]

    def get_queryset(self):
        qs = Gasto.objects.select_related("tienda", "empleado", "sesion_caja")
        qs = scope_qs(self.request, qs)

        user = self.request.user
        if user.rol == 'cajero':
            fecha = self.request.query_params.get("fecha") or str(timezone.now().date())
            return qs.filter(
                tienda_id=user.tienda_id,
                created_at__date=fecha,
                visibilidad='todos',
            ).order_by("-created_at")

        fecha_ini   = self.request.query_params.get("fecha_ini")
        fecha_fin   = self.request.query_params.get("fecha_fin")
        fecha       = self.request.query_params.get("fecha")
        tienda_id   = self.request.query_params.get("tienda_id")
        categoria   = self.request.query_params.get("categoria")
        visibilidad = self.request.query_params.get("visibilidad")

        if fecha_ini and fecha_fin:
            qs = qs.filter(created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin)
        elif fecha:
            qs = qs.filter(created_at__date=fecha)
        else:
            qs = qs.filter(created_at__date=timezone.now().date())

        if tienda_id:   qs = qs.filter(tienda_id=tienda_id)
        if categoria:   qs = qs.filter(categoria__iexact=categoria)
        if visibilidad: qs = qs.filter(visibilidad=visibilidad)

        return qs.order_by("-created_at")

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["request"] = self.request
        return context

    def perform_create(self, serializer):
        tienda_id = self.request.data.get("tienda")
        empresa   = None if es_superadmin(self.request) else get_empresa(self.request)

        sesion_filtro = {"tienda_id": tienda_id, "estado": "abierta"}
        if empresa:
            sesion_filtro["tienda__empresa"] = empresa
        sesion = SesionCaja.objects.filter(**sesion_filtro).first()

        categoria_val       = self.request.data.get("categoria", "").lower().strip()
        visibilidad_enviada = self.request.data.get("visibilidad", "")
        rol                 = self.request.user.rol

        if visibilidad_enviada in ('todos', 'solo_admin'):
            visibilidad = visibilidad_enviada
        elif rol == 'cajero' or categoria_val not in CATEGORIAS_SOLO_ADMIN:
            visibilidad = 'todos'
        else:
            visibilidad = 'solo_admin'

        serializer.save(
            empleado    = self.request.user,
            sesion_caja = sesion,
            visibilidad = visibilidad,
        )


class GastoDetailView(generics.RetrieveDestroyAPIView):
    serializer_class   = GastoSerializer
    permission_classes = [EsAdminOSupervisor]

    def get_queryset(self):
        if es_superadmin(self.request):
            return Gasto.objects.all()
        return Gasto.objects.filter(tienda__empresa=get_empresa(self.request))


# ── Resumen diario ────────────────────────────────────────────
class ResumenDiarioView(APIView):
    permission_classes = [EsAdminSupervisorOCajero]

    def get(self, request):
        fecha     = request.query_params.get("fecha") or timezone.now().date()
        tienda_id = request.query_params.get("tienda_id")

        if request.user.rol == "cajero":
            tienda_id = str(request.user.tienda_id)

        ventas_qs = Venta.objects.filter(estado="completada", created_at__date=fecha)
        gastos_qs = Gasto.objects.filter(created_at__date=fecha)
        dev_qs    = Devolucion.objects.filter(estado="procesada", created_at__date=fecha)

        ventas_qs, gastos_qs, dev_qs = scope_qs(
            request, ventas_qs, gastos_qs, dev_qs, tienda_id=tienda_id
        )

        if request.user.rol == "cajero":
            gastos_qs = gastos_qs.filter(visibilidad='todos')

        total_ventas       = ventas_qs.aggregate(t=Sum("total"))["t"]         or Decimal("0")
        total_gastos       = gastos_qs.aggregate(t=Sum("monto"))["t"]         or Decimal("0")
        total_devoluciones = dev_qs.aggregate(t=Sum("total_devuelto"))["t"]   or Decimal("0")

        por_metodo     = ventas_qs.values("metodo_pago").annotate(total=Sum("total"), cantidad=Count("id"))
        dev_por_metodo = dev_qs.values("metodo_devolucion").annotate(total=Sum("total_devuelto"), cantidad=Count("id"))

        gastos_por_categoria = []
        if request.user.rol != "cajero":
            gastos_por_categoria = [
                {"categoria": g["categoria"], "total": float(g["total"]), "cantidad": g["cantidad"]}
                for g in gastos_qs.values("categoria")
                .annotate(total=Sum("monto"), cantidad=Count("id"))
                .order_by("-total")
            ]

        return Response({
            "fecha":              str(fecha),
            "total_ventas":       float(total_ventas),
            "num_ventas":         ventas_qs.count(),
            "total_gastos":       float(total_gastos),
            "total_devoluciones": float(total_devoluciones),
            "num_devoluciones":   dev_qs.count(),
            "total_neto":         float(total_ventas - total_devoluciones),
            "utilidad_bruta":     float(total_ventas - total_devoluciones - total_gastos),
            "ventas_por_metodo_pago": [
                {"metodo": v["metodo_pago"], "total": float(v["total"]), "cantidad": v["cantidad"]}
                for v in por_metodo
            ],
            "devoluciones_por_metodo": [
                {"metodo": d["metodo_devolucion"], "total": float(d["total"]), "cantidad": d["cantidad"]}
                for d in dev_por_metodo
            ],
            "gastos_por_categoria": gastos_por_categoria,
            "advertencia_cajas_abiertas": _cajas_abiertas(request, tienda_id, str(fecha)),
        })


# ── Resumen mensual ───────────────────────────────────────────
class ResumenMensualView(APIView):
    permission_classes = [EsAdminOSupervisor]

    def get(self, request):
        anio      = request.query_params.get("anio",  str(timezone.now().year))
        mes       = request.query_params.get("mes",   str(timezone.now().month))
        tienda_id = request.query_params.get("tienda_id")

        ventas_qs = Venta.objects.filter(estado="completada", created_at__year=anio, created_at__month=mes)
        gastos_qs = Gasto.objects.filter(created_at__year=anio, created_at__month=mes)
        dev_qs    = Devolucion.objects.filter(estado="procesada", created_at__year=anio, created_at__month=mes)

        ventas_qs, gastos_qs, dev_qs = scope_qs(
            request, ventas_qs, gastos_qs, dev_qs, tienda_id=tienda_id
        )

        total_mes        = ventas_qs.aggregate(t=Sum("total"))["t"]       or Decimal("0")
        gastos_mes       = gastos_qs.aggregate(t=Sum("monto"))["t"]       or Decimal("0")
        devoluciones_mes = dev_qs.aggregate(t=Sum("total_devuelto"))["t"] or Decimal("0")

        por_dia = (
            ventas_qs.annotate(dia=TruncDate("created_at"))
            .values("dia").annotate(total=Sum("total"), cantidad=Count("id")).order_by("dia")
        )
        gastos_por_categoria = (
            gastos_qs.values("categoria").annotate(total=Sum("monto")).order_by("-total")
        )
        gastos_por_dia = (
            gastos_qs.annotate(dia=TruncDate("created_at"))
            .values("dia").annotate(total=Sum("monto")).order_by("dia")
        )
        devoluciones_por_dia = (
            dev_qs.annotate(dia=TruncDate("created_at"))
            .values("dia").annotate(total=Sum("total_devuelto"), cantidad=Count("id")).order_by("dia")
        )

        return Response({
            "anio":               int(anio),
            "mes":                int(mes),
            "total_ventas":       float(total_mes),
            "total_gastos":       float(gastos_mes),
            "total_devoluciones": float(devoluciones_mes),
            "total_neto":         float(total_mes - devoluciones_mes),
            "utilidad_bruta":     float(total_mes - devoluciones_mes - gastos_mes),
            "ventas_por_dia": [
                {"dia": str(v["dia"]), "total": float(v["total"]), "cantidad": v["cantidad"]}
                for v in por_dia
            ],
            "gastos_por_categoria": [
                {"categoria": g["categoria"], "total": float(g["total"])}
                for g in gastos_por_categoria
            ],
            "gastos_por_dia": [
                {"dia": str(g["dia"]), "total": float(g["total"])}
                for g in gastos_por_dia
            ],
            "devoluciones_por_dia": [
                {"dia": str(d["dia"]), "total": float(d["total"]), "cantidad": d["cantidad"]}
                for d in devoluciones_por_dia
            ],
        })


# ── Top productos ─────────────────────────────────────────────
class ProductosMasVendidosView(APIView):
    permission_classes = [EsAdminSupervisorOCajero]

    def get(self, request):
        tienda_id = request.query_params.get("tienda_id")
        fecha_ini = request.query_params.get("fecha_ini")
        fecha_fin = request.query_params.get("fecha_fin")

        if request.user.rol == "cajero":
            tienda_id = str(request.user.tienda_id)

        qs = DetalleVenta.objects.filter(venta__estado="completada")
        qs = scope_qs(request, qs, campo_empresa="venta__tienda__empresa")

        if tienda_id: qs = qs.filter(venta__tienda_id=tienda_id)
        if fecha_ini: qs = qs.filter(venta__created_at__date__gte=fecha_ini)
        if fecha_fin: qs = qs.filter(venta__created_at__date__lte=fecha_fin)

        top = (
            qs.values("producto__nombre")
            .annotate(total_cantidad=Sum("cantidad"), total_ingresos=Sum("subtotal"))
            .order_by("-total_cantidad")[:10]
        )

        return Response([
            {
                "producto":       t["producto__nombre"],
                "total_vendido":  float(t["total_cantidad"]),
                "total_ingresos": float(t["total_ingresos"]),
            }
            for t in top
        ])


# ── Resumen anual ─────────────────────────────────────────────
class ResumenAnualView(APIView):
    permission_classes = [EsAdminOSupervisor]

    def get(self, request):
        anio      = int(request.query_params.get("anio", timezone.now().year))
        tienda_id = request.query_params.get("tienda_id")

        ventas_qs = Venta.objects.filter(estado="completada", created_at__year=anio)
        gastos_qs = Gasto.objects.filter(created_at__year=anio)
        dev_qs    = Devolucion.objects.filter(estado="procesada", created_at__year=anio)

        ventas_qs, gastos_qs, dev_qs = scope_qs(
            request, ventas_qs, gastos_qs, dev_qs, tienda_id=tienda_id
        )

        total_anio        = ventas_qs.aggregate(t=Sum("total"))["t"]       or Decimal("0")
        gastos_anio       = gastos_qs.aggregate(t=Sum("monto"))["t"]       or Decimal("0")
        devoluciones_anio = dev_qs.aggregate(t=Sum("total_devuelto"))["t"] or Decimal("0")

        por_mes_ventas = {
            v["mes"].month: v
            for v in ventas_qs.annotate(mes=TruncMonth("created_at"))
            .values("mes").annotate(total=Sum("total"), cantidad=Count("id"))
        }
        por_mes_gastos = {
            g["mes"].month: float(g["total"])
            for g in gastos_qs.annotate(mes=TruncMonth("created_at"))
            .values("mes").annotate(total=Sum("monto"))
        }
        por_mes_devoluciones = {
            d["mes"].month: float(d["total"])
            for d in dev_qs.annotate(mes=TruncMonth("created_at"))
            .values("mes").annotate(total=Sum("total_devuelto"))
        }

        meses = []
        for m in range(1, 13):
            ventas_m = float(por_mes_ventas.get(m, {}).get("total", 0) or 0)
            gastos_m = por_mes_gastos.get(m, 0)
            dev_m    = por_mes_devoluciones.get(m, 0)
            meses.append({
                "mes":          m,
                "nombre":       MESES[m],
                "ventas":       ventas_m,
                "devoluciones": dev_m,
                "neto":         ventas_m - dev_m,
                "gastos":       gastos_m,
                "utilidad":     ventas_m - dev_m - gastos_m,
                "cantidad":     por_mes_ventas.get(m, {}).get("cantidad", 0) or 0,
            })

        return Response({
            "anio":               anio,
            "total_ventas":       float(total_anio),
            "total_gastos":       float(gastos_anio),
            "total_devoluciones": float(devoluciones_anio),
            "total_neto":         float(total_anio - devoluciones_anio),
            "utilidad_bruta":     float(total_anio - devoluciones_anio - gastos_anio),
            "meses":              meses,
        })


# ── Gastos por rango ──────────────────────────────────────────
class GastosResumenRangoView(APIView):
    permission_classes = [EsAdminOSupervisor]

    def get(self, request):
        fecha_ini = request.query_params.get("fecha_ini")
        fecha_fin = request.query_params.get("fecha_fin")
        tienda_id = request.query_params.get("tienda_id")
        categoria = request.query_params.get("categoria")

        if not fecha_ini or not fecha_fin:
            return Response({"error": "fecha_ini y fecha_fin son requeridos"}, status=400)

        qs = Gasto.objects.filter(created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin)
        qs = scope_qs(request, qs, tienda_id=tienda_id)

        if categoria: qs = qs.filter(categoria__iexact=categoria)

        total = qs.aggregate(t=Sum("monto"))["t"] or Decimal("0")

        por_categoria = (
            qs.values("categoria")
            .annotate(total=Sum("monto"), cantidad=Count("id"))
            .order_by("-total")
        )
        por_dia = (
            qs.annotate(dia=TruncDate("created_at"))
            .values("dia").annotate(total=Sum("monto"), cantidad=Count("id")).order_by("dia")
        )

        return Response({
            "fecha_ini": fecha_ini,
            "fecha_fin": fecha_fin,
            "total":     float(total),
            "cantidad":  qs.count(),
            "por_categoria": [
                {"categoria": g["categoria"], "total": float(g["total"]), "cantidad": g["cantidad"]}
                for g in por_categoria
            ],
            "por_dia": [
                {"dia": str(g["dia"]), "total": float(g["total"]), "cantidad": g["cantidad"]}
                for g in por_dia
            ],
        })


# ── Estado de Resultados (P&L) ────────────────────────────────
class EstadoResultadosView(APIView):
    """
    Devuelve un estado de resultados completo para el período indicado.
    Parámetros: fecha_ini, fecha_fin (ISO), tienda_id (opcional).
    Si no se envían fechas, usa el mes actual.
    """
    permission_classes = [EsAdminOSupervisor]

    def get(self, request):
        today     = timezone.now().date()
        fecha_ini = request.query_params.get("fecha_ini") or today.replace(day=1).isoformat()
        fecha_fin = request.query_params.get("fecha_fin") or today.isoformat()
        tienda_id = request.query_params.get("tienda_id")

        # ── Querysets base ────────────────────────────────────
        ventas_qs = Venta.objects.filter(
            estado="completada",
            created_at__date__gte=fecha_ini,
            created_at__date__lte=fecha_fin,
        )
        dev_qs = Devolucion.objects.filter(
            estado="procesada",
            created_at__date__gte=fecha_ini,
            created_at__date__lte=fecha_fin,
        )
        gastos_qs = Gasto.objects.filter(
            created_at__date__gte=fecha_ini,
            created_at__date__lte=fecha_fin,
        )
        detalles_qs = DetalleVenta.objects.filter(
            venta__estado="completada",
            venta__created_at__date__gte=fecha_ini,
            venta__created_at__date__lte=fecha_fin,
        )
        mov_qs = MovimientoInventario.objects.filter(
            created_at__date__gte=fecha_ini,
            created_at__date__lte=fecha_fin,
        )

        # ── Scope por empresa/tienda ──────────────────────────
        ventas_qs, dev_qs, gastos_qs, mov_qs = scope_qs(
            request, ventas_qs, dev_qs, gastos_qs, mov_qs,
            tienda_id=tienda_id,
        )
        # DetalleVenta no tiene tienda directo, se maneja aparte
        detalles_qs = scope_qs(
            request, detalles_qs, campo_empresa="venta__tienda__empresa"
        )
        if tienda_id:
            detalles_qs = detalles_qs.filter(venta__tienda_id=tienda_id)

        # ── Ingresos ──────────────────────────────────────────
        agg_v = ventas_qs.aggregate(
            bruto=Sum("total"),
            descuentos=Sum("descuento_total"),
            impuestos=Sum("impuesto_total"),
            num=Count("id"),
        )
        ventas_brutas    = agg_v["bruto"]     or Decimal("0")
        total_descuentos = agg_v["descuentos"] or Decimal("0")
        total_impuestos  = agg_v["impuestos"]  or Decimal("0")
        num_ventas       = agg_v["num"]        or 0

        agg_d = dev_qs.aggregate(total=Sum("total_devuelto"), num=Count("id"))
        total_devoluciones = agg_d["total"] or Decimal("0")
        num_devoluciones   = agg_d["num"]   or 0

        ingresos_netos = ventas_brutas - total_devoluciones

        # ── Costo de ventas (COGS) ────────────────────────────
        # Usa costo_unitario guardado en el momento de la venta;
        # cae en precio_compra actual para registros históricos.
        _dec = DecimalField(max_digits=15, decimal_places=2)
        cogs = detalles_qs.aggregate(
            total=Sum(
                F("cantidad") * Coalesce(
                    F("costo_unitario"), F("producto__precio_compra"),
                    output_field=_dec,
                ),
                output_field=_dec,
            )
        )["total"] or Decimal("0")

        margen_bruto     = ingresos_netos - cogs
        margen_bruto_pct = (
            round(float(margen_bruto) / float(ingresos_netos) * 100, 2)
            if ingresos_netos > 0 else 0.0
        )

        # ── Gastos operativos ─────────────────────────────────
        total_gastos = gastos_qs.aggregate(t=Sum("monto"))["t"] or Decimal("0")
        gastos_detalle = [
            {"categoria": g["categoria"], "total": float(g["total"]), "cantidad": g["cantidad"]}
            for g in gastos_qs.values("categoria")
            .annotate(total=Sum("monto"), cantidad=Count("id"))
            .order_by("-total")
        ]

        # ── Averías ───────────────────────────────────────────
        _val = lambda qs: qs.aggregate(
            v=Sum(F("cantidad") * F("producto__precio_compra"), output_field=_dec)
        )["v"] or Decimal("0")

        perdidas_brutas  = _val(mov_qs.filter(tipo="dano"))
        valor_recuperado = _val(mov_qs.filter(referencia_tipo="recuperacion_averia"))
        perdida_neta_averias = max(Decimal("0"), perdidas_brutas - valor_recuperado)

        # ── Resultado final ───────────────────────────────────
        utilidad_operativa = margen_bruto - total_gastos - perdida_neta_averias
        utilidad_operativa_pct = (
            round(float(utilidad_operativa) / float(ingresos_netos) * 100, 2)
            if ingresos_netos > 0 else 0.0
        )

        return Response({
            "periodo": {
                "desde":     str(fecha_ini),
                "hasta":     str(fecha_fin),
                "tienda_id": tienda_id,
            },
            "ingresos": {
                "ventas_brutas":      float(ventas_brutas),
                "menos_descuentos":   float(total_descuentos),
                "menos_devoluciones": float(total_devoluciones),
                "ingresos_netos":     float(ingresos_netos),
                "impuestos_cobrados": float(total_impuestos),
                "num_ventas":         num_ventas,
                "num_devoluciones":   num_devoluciones,
            },
            "costo_ventas":    float(cogs),
            "margen_bruto":    float(margen_bruto),
            "margen_bruto_pct": margen_bruto_pct,
            "gastos_operativos": {
                "total":   float(total_gastos),
                "detalle": gastos_detalle,
            },
            "averias": {
                "perdidas_brutas":  float(perdidas_brutas),
                "valor_recuperado": float(valor_recuperado),
                "perdida_neta":     float(perdida_neta_averias),
            },
            "utilidad_operativa":          float(utilidad_operativa),
            "utilidad_operativa_pct":      utilidad_operativa_pct,
            "advertencia_cajas_abiertas":  _cajas_abiertas(request, tienda_id, fecha_fin),
        })


# ── Comparativo entre tiendas ─────────────────────────────────
class ComparativoTiendasView(APIView):
    """
    Muestra métricas P&L side-by-side para cada tienda de la empresa.
    Incluye fila de totales consolidados al final.
    Parámetros: fecha_ini, fecha_fin (ISO). Por defecto mes actual.
    """
    permission_classes = [EsAdminOSupervisor]

    def get(self, request):
        today     = timezone.now().date()
        fecha_ini = request.query_params.get("fecha_ini") or today.replace(day=1).isoformat()
        fecha_fin = request.query_params.get("fecha_fin") or today.isoformat()

        ventas_qs = Venta.objects.filter(
            estado="completada",
            created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin,
        )
        dev_qs = Devolucion.objects.filter(
            estado="procesada",
            created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin,
        )
        gastos_qs = Gasto.objects.filter(
            created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin,
        )
        detalles_qs = DetalleVenta.objects.filter(
            venta__estado="completada",
            venta__created_at__date__gte=fecha_ini, venta__created_at__date__lte=fecha_fin,
        )
        mov_qs = MovimientoInventario.objects.filter(
            created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin,
        )

        # Sin tienda_id: ver todas las tiendas de la empresa
        ventas_qs, dev_qs, gastos_qs, mov_qs = scope_qs(
            request, ventas_qs, dev_qs, gastos_qs, mov_qs,
        )
        detalles_qs = scope_qs(request, detalles_qs, campo_empresa="venta__tienda__empresa")

        _dec = DecimalField(max_digits=15, decimal_places=2)

        ventas_map = {
            r["tienda_id"]: r
            for r in ventas_qs.values("tienda_id", "tienda__nombre").annotate(
                ventas_brutas=Sum("total"),
                total_descuentos=Sum("descuento_total"),
                num_ventas=Count("id"),
            )
        }
        dev_map = {
            r["tienda_id"]: r
            for r in dev_qs.values("tienda_id").annotate(
                total_devuelto=Sum("total_devuelto"),
                num_devoluciones=Count("id"),
            )
        }
        gastos_map = {
            r["tienda_id"]: float(r["total"] or 0)
            for r in gastos_qs.values("tienda_id").annotate(total=Sum("monto"))
        }
        cogs_map = {
            r["venta__tienda_id"]: float(r["cogs"] or 0)
            for r in detalles_qs.values("venta__tienda_id").annotate(
                cogs=Sum(
                    F("cantidad") * Coalesce(
                        F("costo_unitario"), F("producto__precio_compra"), output_field=_dec,
                    ),
                    output_field=_dec,
                )
            )
        }
        danos_map = {
            r["tienda_id"]: float(r["v"] or 0)
            for r in mov_qs.filter(tipo="dano").values("tienda_id").annotate(
                v=Sum(F("cantidad") * F("producto__precio_compra"), output_field=_dec)
            )
        }
        recuperado_map = {
            r["tienda_id"]: float(r["v"] or 0)
            for r in mov_qs.filter(referencia_tipo="recuperacion_averia").values("tienda_id").annotate(
                v=Sum(F("cantidad") * F("producto__precio_compra"), output_field=_dec)
            )
        }

        tiendas = []
        for tid, vd in ventas_map.items():
            ventas_brutas   = float(vd["ventas_brutas"] or 0)
            devuelto        = float(dev_map.get(tid, {}).get("total_devuelto") or 0)
            ingresos_netos  = ventas_brutas - devuelto
            cogs            = cogs_map.get(tid, 0)
            margen_bruto    = ingresos_netos - cogs
            total_gastos    = gastos_map.get(tid, 0)
            perdida_averias = max(0, danos_map.get(tid, 0) - recuperado_map.get(tid, 0))
            utilidad        = margen_bruto - total_gastos - perdida_averias

            tiendas.append({
                "tienda_id":          tid,
                "tienda_nombre":      vd["tienda__nombre"],
                "ventas_brutas":      ventas_brutas,
                "devoluciones":       devuelto,
                "ingresos_netos":     ingresos_netos,
                "costo_ventas":       cogs,
                "margen_bruto":       margen_bruto,
                "margen_bruto_pct":   round(margen_bruto / ingresos_netos * 100, 2) if ingresos_netos > 0 else 0.0,
                "gastos":             total_gastos,
                "perdida_averias":    perdida_averias,
                "utilidad_operativa": utilidad,
                "num_ventas":         vd["num_ventas"] or 0,
                "num_devoluciones":   dev_map.get(tid, {}).get("num_devoluciones") or 0,
            })

        tiendas.sort(key=lambda x: x["utilidad_operativa"], reverse=True)

        def _sum(key):
            return round(sum(t[key] for t in tiendas), 2)

        ingresos_total = _sum("ingresos_netos")
        margen_total   = _sum("margen_bruto")
        totales = {
            "ventas_brutas":      _sum("ventas_brutas"),
            "devoluciones":       _sum("devoluciones"),
            "ingresos_netos":     ingresos_total,
            "costo_ventas":       _sum("costo_ventas"),
            "margen_bruto":       margen_total,
            "margen_bruto_pct":   round(margen_total / ingresos_total * 100, 2) if ingresos_total > 0 else 0.0,
            "gastos":             _sum("gastos"),
            "perdida_averias":    _sum("perdida_averias"),
            "utilidad_operativa": _sum("utilidad_operativa"),
            "num_ventas":         _sum("num_ventas"),
        }

        return Response({
            "periodo": {"desde": str(fecha_ini), "hasta": str(fecha_fin)},
            "tiendas": tiendas,
            "totales": totales,
        })


# ── Ventas por empleado ───────────────────────────────────────
class VentasPorEmpleadoView(APIView):
    """
    Ranking de ventas por empleado/cajero en el período.
    Parámetros: fecha_ini, fecha_fin, tienda_id (opcional).
    """
    permission_classes = [EsAdminOSupervisor]

    def get(self, request):
        today     = timezone.now().date()
        fecha_ini = request.query_params.get("fecha_ini") or today.replace(day=1).isoformat()
        fecha_fin = request.query_params.get("fecha_fin") or today.isoformat()
        tienda_id = request.query_params.get("tienda_id")

        ventas_qs = Venta.objects.filter(
            estado="completada",
            created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin,
        )
        ventas_qs = scope_qs(request, ventas_qs, tienda_id=tienda_id)

        data = (
            ventas_qs
            .values("empleado_id", "empleado__nombre", "empleado__apellido")
            .annotate(
                num_ventas=Count("id"),
                total_ventas=Sum("total"),
                total_descuentos=Sum("descuento_total"),
            )
            .order_by("-total_ventas")
        )

        empleados = []
        for r in data:
            total  = float(r["total_ventas"] or 0)
            num    = r["num_ventas"] or 0
            nombre = f"{r['empleado__nombre'] or ''} {r['empleado__apellido'] or ''}".strip()
            empleados.append({
                "empleado_id":      r["empleado_id"],
                "nombre":           nombre or "Sin asignar",
                "num_ventas":       num,
                "total_ventas":     total,
                "total_descuentos": float(r["total_descuentos"] or 0),
                "promedio_venta":   round(total / num, 2) if num else 0.0,
            })

        return Response({
            "periodo":   {"desde": str(fecha_ini), "hasta": str(fecha_fin), "tienda_id": tienda_id},
            "empleados": empleados,
        })


# ── Punto de Equilibrio ───────────────────────────────────────
class PuntoEquilibrioView(APIView):
    """
    Calcula el punto de equilibrio para el período indicado.

    Fórmula:
      Margen de contribución % = (Ingresos netos - COGS - Gastos variables) / Ingresos netos
      Punto de equilibrio      = Gastos fijos / Margen de contribución %

    El punto de equilibrio indica cuánto necesita vender la tienda/empresa
    para cubrir todos sus costos sin generar pérdidas.

    Parámetros: fecha_ini, fecha_fin (ISO), tienda_id (opcional).
    """
    permission_classes = [EsAdminOSupervisor]

    def get(self, request):
        today     = timezone.now().date()
        fecha_ini = request.query_params.get("fecha_ini") or today.replace(day=1).isoformat()
        fecha_fin = request.query_params.get("fecha_fin") or today.isoformat()
        tienda_id = request.query_params.get("tienda_id")

        ventas_qs = Venta.objects.filter(
            estado="completada",
            created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin,
        )
        dev_qs = Devolucion.objects.filter(
            estado="procesada",
            created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin,
        )
        gastos_qs = Gasto.objects.filter(
            created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin,
        )
        detalles_qs = DetalleVenta.objects.filter(
            venta__estado="completada",
            venta__created_at__date__gte=fecha_ini, venta__created_at__date__lte=fecha_fin,
        )

        ventas_qs, dev_qs, gastos_qs = scope_qs(
            request, ventas_qs, dev_qs, gastos_qs, tienda_id=tienda_id,
        )
        detalles_qs = scope_qs(request, detalles_qs, campo_empresa="venta__tienda__empresa")
        if tienda_id:
            detalles_qs = detalles_qs.filter(venta__tienda_id=tienda_id)

        _dec = DecimalField(max_digits=15, decimal_places=2)

        # ── Ingresos netos ────────────────────────────────────
        ventas_brutas      = ventas_qs.aggregate(t=Sum("total"))["t"]          or Decimal("0")
        total_devoluciones = dev_qs.aggregate(t=Sum("total_devuelto"))["t"]    or Decimal("0")
        ingresos_netos     = ventas_brutas - total_devoluciones

        # ── COGS ──────────────────────────────────────────────
        cogs = detalles_qs.aggregate(
            total=Sum(
                F("cantidad") * Coalesce(
                    F("costo_unitario"), F("producto__precio_compra"), output_field=_dec,
                ),
                output_field=_dec,
            )
        )["total"] or Decimal("0")

        # ── Gastos fijos y variables ──────────────────────────
        gastos_agg = gastos_qs.values("tipo_gasto").annotate(total=Sum("monto"))
        gastos_fijos     = Decimal("0")
        gastos_variables = Decimal("0")
        for g in gastos_agg:
            if g["tipo_gasto"] == "variable":
                gastos_variables += g["total"] or Decimal("0")
            else:
                gastos_fijos += g["total"] or Decimal("0")

        detalle_fijos = [
            {"categoria": g["categoria"], "total": float(g["total"])}
            for g in gastos_qs.filter(tipo_gasto="fijo")
            .values("categoria").annotate(total=Sum("monto")).order_by("-total")
        ]
        detalle_variables = [
            {"categoria": g["categoria"], "total": float(g["total"])}
            for g in gastos_qs.filter(tipo_gasto="variable")
            .values("categoria").annotate(total=Sum("monto")).order_by("-total")
        ]

        # ── Margen de contribución ────────────────────────────
        margen_contribucion = ingresos_netos - cogs - gastos_variables
        margen_contribucion_pct = (
            round(float(margen_contribucion) / float(ingresos_netos) * 100, 4)
            if ingresos_netos > 0 else 0.0
        )

        # ── Punto de equilibrio ───────────────────────────────
        if margen_contribucion_pct > 0:
            pe_ingresos = float(gastos_fijos) / (margen_contribucion_pct / 100)
        else:
            pe_ingresos = None  # no calculable: costos superan ingresos

        alcanzado  = pe_ingresos is not None and float(ingresos_netos) >= pe_ingresos
        excedente  = (
            round(float(ingresos_netos) - pe_ingresos, 2)
            if pe_ingresos is not None else None
        )

        return Response({
            "periodo": {
                "desde":     str(fecha_ini),
                "hasta":     str(fecha_fin),
                "tienda_id": tienda_id,
            },
            "ingresos_netos":           float(ingresos_netos),
            "costo_ventas":             float(cogs),
            "gastos_fijos":             float(gastos_fijos),
            "gastos_variables":         float(gastos_variables),
            "margen_contribucion":      float(margen_contribucion),
            "margen_contribucion_pct":  margen_contribucion_pct,
            "punto_equilibrio_ingresos": round(pe_ingresos, 2) if pe_ingresos is not None else None,
            "punto_equilibrio_alcanzado": alcanzado,
            "excedente_deficit":        excedente,
            "detalle_gastos_fijos":     detalle_fijos,
            "detalle_gastos_variables": detalle_variables,
        })


# ── Flujo de Caja ─────────────────────────────────────────────
class FlujoCajaView(APIView):
    """
    Flujo de caja real basado en sesiones de caja cerradas/abiertas.
    Cada sesión muestra entradas (ventas + abonos), salidas (gastos +
    devoluciones) y la diferencia al cierre.
    Parámetros: fecha_ini, fecha_fin (ISO), tienda_id (opcional).
    """
    permission_classes = [EsAdminOSupervisor]

    def get(self, request):
        today     = timezone.now().date()
        fecha_ini = request.query_params.get("fecha_ini") or today.replace(day=1).isoformat()
        fecha_fin = request.query_params.get("fecha_fin") or today.isoformat()
        tienda_id = request.query_params.get("tienda_id")

        sesiones_qs = SesionCaja.objects.filter(
            fecha_apertura__date__gte=fecha_ini,
            fecha_apertura__date__lte=fecha_fin,
        ).select_related("tienda", "empleado").order_by("fecha_apertura")

        sesiones_qs = scope_qs(
            request, sesiones_qs,
            campo_empresa="tienda__empresa", tienda_id=tienda_id,
        )

        sesion_ids = list(sesiones_qs.values_list("id", flat=True))

        # ── Ventas por sesión y método de pago ────────────────
        ventas_raw = (
            Venta.objects.filter(sesion_caja_id__in=sesion_ids, estado="completada")
            .values("sesion_caja_id", "metodo_pago")
            .annotate(total=Sum("total"))
        )
        ventas_map = {}
        for v in ventas_raw:
            sid = v["sesion_caja_id"]
            ventas_map.setdefault(sid, {})[v["metodo_pago"]] = float(v["total"] or 0)

        # ── Gastos por sesión ─────────────────────────────────
        gastos_map = {
            g["sesion_caja_id"]: float(g["total"] or 0)
            for g in Gasto.objects.filter(sesion_caja_id__in=sesion_ids)
            .values("sesion_caja_id").annotate(total=Sum("monto"))
        }

        # ── Abonos por sesión (MovimientoCaja tipo abono_separado) ──
        abonos_map = {
            a["sesion_id"]: float(a["total"] or 0)
            for a in MovimientoCaja.objects.filter(
                sesion_id__in=sesion_ids, tipo="abono_separado"
            ).values("sesion_id").annotate(total=Sum("monto"))
        }

        # ── Devoluciones por sesión ───────────────────────────
        dev_map = {
            d["venta__sesion_caja_id"]: float(d["total"] or 0)
            for d in Devolucion.objects.filter(
                venta__sesion_caja_id__in=sesion_ids, estado="procesada"
            ).values("venta__sesion_caja_id").annotate(total=Sum("total_devuelto"))
        }

        # ── Construir resultado sesión a sesión ───────────────
        sesiones = []
        for s in sesiones_qs:
            vm              = ventas_map.get(s.id, {})
            v_efectivo      = vm.get("efectivo", 0)
            v_tarjeta       = vm.get("tarjeta", 0)
            v_transferencia = vm.get("transferencia", 0)
            v_mixto         = vm.get("mixto", 0)
            abonos          = abonos_map.get(s.id, 0)
            gastos          = gastos_map.get(s.id, 0)
            devoluciones    = dev_map.get(s.id, 0)

            total_entradas = v_efectivo + v_tarjeta + v_transferencia + v_mixto + abonos
            total_salidas  = gastos + devoluciones

            sesiones.append({
                "sesion_id":      s.id,
                "fecha":          s.fecha_apertura.date().isoformat(),
                "tienda_id":      s.tienda_id,
                "tienda_nombre":  s.tienda.nombre,
                "empleado": (
                    f"{s.empleado.nombre} {s.empleado.apellido}" if s.empleado else None
                ),
                "monto_inicial":  float(s.monto_inicial),
                "entradas": {
                    "ventas_efectivo":      v_efectivo,
                    "ventas_tarjeta":       v_tarjeta,
                    "ventas_transferencia": v_transferencia,
                    "ventas_mixto":         v_mixto,
                    "abonos":               abonos,
                    "total":                round(total_entradas, 2),
                },
                "salidas": {
                    "gastos":       gastos,
                    "devoluciones": devoluciones,
                    "total":        round(total_salidas, 2),
                },
                "flujo_sesion":        round(total_entradas - total_salidas, 2),
                "monto_final_sistema": float(s.monto_final_sistema) if s.monto_final_sistema is not None else None,
                "monto_final_real":    float(s.monto_final_real)    if s.monto_final_real    is not None else None,
                "diferencia":          float(s.diferencia)          if s.diferencia          is not None else None,
                "estado":              s.estado,
            })

        total_entradas   = round(sum(s["entradas"]["total"] for s in sesiones), 2)
        total_salidas    = round(sum(s["salidas"]["total"]  for s in sesiones), 2)
        total_dif        = round(sum(s["diferencia"] or 0   for s in sesiones), 2)

        return Response({
            "periodo": {"desde": str(fecha_ini), "hasta": str(fecha_fin), "tienda_id": tienda_id},
            "resumen": {
                "total_entradas":    total_entradas,
                "total_salidas":     total_salidas,
                "flujo_neto":        round(total_entradas - total_salidas, 2),
                "total_diferencias": total_dif,
                "num_sesiones":      len(sesiones),
            },
            "sesiones":                   sesiones,
            "advertencia_cajas_abiertas": _cajas_abiertas(request, tienda_id, fecha_fin),
        })


# ── Exportación a Excel ───────────────────────────────────────
# Helpers de estilo
_NAVY    = "1E3A5F"
_BLUE    = "2E6DA4"
_GRAY    = "F2F2F2"
_GREEN   = "1E7E34"
_RED     = "B22222"
_FMT_CUR = '#,##0.00'
_FMT_PCT = '0.00%'


def _hdr(ws, row, col, value, bold=True, bg=None, color="FFFFFF", size=11, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color, size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell


def _val(ws, row, col, value, bold=False, color="000000", fmt=_FMT_CUR, bg=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


def _build_estado_resultados_sheet(ws, data):
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18

    r = 1
    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "ESTADO DE RESULTADOS", bold=True, bg=_NAVY, size=13)
    r += 1
    ws.merge_cells(f"A{r}:B{r}")
    periodo = f"{data['periodo']['desde']}  →  {data['periodo']['hasta']}"
    _hdr(ws, r, 1, periodo, bold=False, bg=_NAVY, color="CCDDEE", size=10)
    r += 2

    # Ingresos
    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "INGRESOS", bold=True, bg=_BLUE, size=11)
    r += 1
    ing = data["ingresos"]
    rows_ing = [
        ("Ventas brutas",          ing["ventas_brutas"],      False),
        ("(-) Descuentos",        -ing["menos_descuentos"],   False),
        ("(-) Devoluciones",      -ing["menos_devoluciones"], False),
        (">> Ingresos netos",      ing["ingresos_netos"],     True),
        ("Impuestos cobrados",     ing["impuestos_cobrados"], False),
    ]
    for label, val, bold in rows_ing:
        _hdr(ws, r, 1, label, bold=bold, bg=_GRAY if bold else None)
        color = _GREEN if val >= 0 else _RED
        _val(ws, r, 2, val, bold=bold, color=color, bg=_GRAY if bold else None)
        r += 1

    r += 1
    # Costo de ventas
    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "COSTO DE VENTAS", bold=True, bg=_BLUE, size=11)
    r += 1
    _hdr(ws, r, 1, "Costo de ventas (COGS)")
    _val(ws, r, 2, -data["costo_ventas"], color=_RED)
    r += 1
    _hdr(ws, r, 1, ">> Margen bruto", bold=True, bg=_GRAY)
    _val(ws, r, 2, data["margen_bruto"], bold=True,
         color=_GREEN if data["margen_bruto"] >= 0 else _RED, bg=_GRAY)
    r += 1
    _hdr(ws, r, 1, f"  Margen bruto %", bold=False, color="555555")
    _val(ws, r, 2, data["margen_bruto_pct"] / 100, fmt=_FMT_PCT, color="555555")
    r += 2

    # Gastos operativos
    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "GASTOS OPERATIVOS", bold=True, bg=_BLUE, size=11)
    r += 1
    for g in data["gastos_operativos"]["detalle"]:
        cat = g["categoria"] or "Sin categoría"
        _hdr(ws, r, 1, f"  {cat.capitalize()}")
        _val(ws, r, 2, -g["total"], color=_RED)
        r += 1
    _hdr(ws, r, 1, ">> Total gastos operativos", bold=True, bg=_GRAY)
    _val(ws, r, 2, -data["gastos_operativos"]["total"], bold=True, color=_RED, bg=_GRAY)
    r += 2

    # Averías
    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "AVERÍAS / DAÑOS", bold=True, bg=_BLUE, size=11)
    r += 1
    av = data["averias"]
    _hdr(ws, r, 1, "  Pérdidas brutas por daños")
    _val(ws, r, 2, -av["perdidas_brutas"], color=_RED)
    r += 1
    _hdr(ws, r, 1, "  (+) Valor recuperado")
    _val(ws, r, 2, av["valor_recuperado"], color=_GREEN)
    r += 1
    _hdr(ws, r, 1, ">> Pérdida neta por averías", bold=True, bg=_GRAY)
    _val(ws, r, 2, -av["perdida_neta"], bold=True, color=_RED, bg=_GRAY)
    r += 2

    # Resultado
    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "RESULTADO", bold=True, bg=_NAVY, size=11)
    r += 1
    uo = data["utilidad_operativa"]
    _hdr(ws, r, 1, "UTILIDAD OPERATIVA", bold=True, bg=_NAVY, color="FFFFFF", size=12)
    _val(ws, r, 2, uo, bold=True,
         color=_GREEN if uo >= 0 else _RED, bg=_NAVY, fmt=_FMT_CUR)
    ws.cell(row=r, column=2).font = Font(bold=True, color=_GREEN if uo >= 0 else _RED, size=12)
    r += 1
    _hdr(ws, r, 1, "  Margen operativo %", bold=False, color="AAAAAA")
    _val(ws, r, 2, data["utilidad_operativa_pct"] / 100, fmt=_FMT_PCT, color="AAAAAA")
    r += 1
    _hdr(ws, r, 1, f"  N° ventas: {data['ingresos']['num_ventas']}  |  "
                   f"N° devoluciones: {data['ingresos']['num_devoluciones']}")


def _build_flujo_caja_sheet(ws, data):
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16

    # Título
    ws.merge_cells("A1:L1")
    _hdr(ws, 1, 1, "FLUJO DE CAJA", bold=True, bg=_NAVY, size=13)
    ws.merge_cells("A2:L2")
    periodo = f"{data['periodo']['desde']}  →  {data['periodo']['hasta']}"
    _hdr(ws, 2, 1, periodo, bold=False, bg=_NAVY, color="CCDDEE", size=10)

    # Encabezados
    cols = [
        "Fecha", "Tienda", "Empleado", "Saldo inicial",
        "V. Efectivo", "V. Tarjeta", "V. Transferencia", "V. Mixto", "Abonos",
        "Gastos", "Devoluciones", "Flujo sesión",
        "Cierre sistema", "Cierre real", "Diferencia", "Estado",
    ]
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = 14 if i > 3 else (10 if i == 1 else 18)
        _hdr(ws, 3, i, c, bold=True, bg=_BLUE, size=10)

    r = 4
    for s in data["sesiones"]:
        ent = s["entradas"]
        sal = s["salidas"]
        flujo = s["flujo_sesion"]
        row_vals = [
            s["fecha"], s["tienda_nombre"], s["empleado"] or "",
            s["monto_inicial"],
            ent["ventas_efectivo"], ent["ventas_tarjeta"],
            ent["ventas_transferencia"], ent["ventas_mixto"], ent["abonos"],
            -sal["gastos"], -sal["devoluciones"],
            flujo,
            s["monto_final_sistema"], s["monto_final_real"], s["diferencia"],
            s["estado"],
        ]
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            if c_idx >= 4 and isinstance(val, (int, float)) and val is not None:
                cell.number_format = _FMT_CUR
                cell.alignment = Alignment(horizontal="right")
                if c_idx == 12:  # flujo sesión
                    cell.font = Font(color=_GREEN if flujo >= 0 else _RED, bold=True)
                if c_idx == 15 and val is not None:  # diferencia
                    cell.font = Font(color=_RED if val < 0 else "000000")
        r += 1

    # Fila de totales
    res = data["resumen"]
    ws.merge_cells(f"A{r}:C{r}")
    _hdr(ws, r, 1, "TOTALES", bold=True, bg=_GRAY)
    totales = [
        None, None, None, None,
        None, None, None, None, None,
        -res["total_salidas"], None,
        res["flujo_neto"],
        None, None, res["total_diferencias"], None,
    ]
    for c_idx, val in enumerate(totales, 1):
        if val is not None:
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.number_format = _FMT_CUR
            cell.font = Font(bold=True, color=_GREEN if val >= 0 else _RED)
            cell.fill = PatternFill("solid", fgColor=_GRAY)
            cell.alignment = Alignment(horizontal="right")


class ExportarContabilidadView(APIView):
    """
    Exporta reportes de contabilidad a Excel.
    Parámetro requerido: tipo = "estado-resultados" | "flujo-caja"
    Mismos parámetros de fecha que las vistas JSON correspondientes.
    """
    permission_classes = [EsAdminOSupervisor]

    def get(self, request):
        import traceback
        tipo      = request.query_params.get("tipo", "estado-resultados")
        today     = timezone.now().date()
        fecha_ini = request.query_params.get("fecha_ini") or today.replace(day=1).isoformat()
        fecha_fin = request.query_params.get("fecha_fin") or today.isoformat()
        tienda_id = request.query_params.get("tienda_id")

        cajas = _cajas_abiertas(request, tienda_id, fecha_fin)
        if cajas:
            nombres = ", ".join(c["tienda"] for c in cajas)
            return Response({
                "error": (
                    f"Hay cajas abiertas en: {nombres}. "
                    "Ciérralas antes de exportar el reporte para evitar datos incompletos."
                ),
                "cajas_abiertas": cajas,
            }, status=409)

        try:
            wb = Workbook()
            ws = wb.active

            if tipo == "estado-resultados":
                ws.title = "Estado de Resultados"
                data = self._get_estado_resultados(request, fecha_ini, fecha_fin, tienda_id)
                _build_estado_resultados_sheet(ws, data)
                nombre = f"estado_resultados_{fecha_ini}_{fecha_fin}.xlsx"

            elif tipo == "flujo-caja":
                ws.title = "Flujo de Caja"
                data = self._get_flujo_caja(request, fecha_ini, fecha_fin, tienda_id)
                _build_flujo_caja_sheet(ws, data)
                nombre = f"flujo_caja_{fecha_ini}_{fecha_fin}.xlsx"

            else:
                return Response({"error": "tipo inválido. Usa: estado-resultados | flujo-caja"}, status=400)

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{nombre}"'
            wb.save(response)
            return response

        except Exception as exc:
            return Response({"error": str(exc), "detalle": traceback.format_exc()}, status=500)


    # ── Lógica reutilizada de EstadoResultadosView ────────────
    def _get_estado_resultados(self, request, fecha_ini, fecha_fin, tienda_id):
        ventas_qs = Venta.objects.filter(
            estado="completada",
            created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin,
        )
        dev_qs = Devolucion.objects.filter(
            estado="procesada",
            created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin,
        )
        gastos_qs = Gasto.objects.filter(
            created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin,
        )
        detalles_qs = DetalleVenta.objects.filter(
            venta__estado="completada",
            venta__created_at__date__gte=fecha_ini, venta__created_at__date__lte=fecha_fin,
        )
        mov_qs = MovimientoInventario.objects.filter(
            created_at__date__gte=fecha_ini, created_at__date__lte=fecha_fin,
        )
        ventas_qs, dev_qs, gastos_qs, mov_qs = scope_qs(
            request, ventas_qs, dev_qs, gastos_qs, mov_qs, tienda_id=tienda_id,
        )
        detalles_qs = scope_qs(request, detalles_qs, campo_empresa="venta__tienda__empresa")
        if tienda_id:
            detalles_qs = detalles_qs.filter(venta__tienda_id=tienda_id)

        _dec = DecimalField(max_digits=15, decimal_places=2)

        agg_v = ventas_qs.aggregate(
            bruto=Sum("total"), descuentos=Sum("descuento_total"),
            impuestos=Sum("impuesto_total"), num=Count("id"),
        )
        ventas_brutas    = agg_v["bruto"]     or Decimal("0")
        total_descuentos = agg_v["descuentos"] or Decimal("0")
        total_impuestos  = agg_v["impuestos"]  or Decimal("0")
        num_ventas       = agg_v["num"]        or 0

        agg_d = dev_qs.aggregate(total=Sum("total_devuelto"), num=Count("id"))
        total_devoluciones = agg_d["total"] or Decimal("0")
        num_devoluciones   = agg_d["num"]   or 0
        ingresos_netos     = ventas_brutas - total_devoluciones

        cogs = detalles_qs.aggregate(
            total=Sum(
                F("cantidad") * Coalesce(F("costo_unitario"), F("producto__precio_compra"), output_field=_dec),
                output_field=_dec,
            )
        )["total"] or Decimal("0")

        margen_bruto     = ingresos_netos - cogs
        margen_bruto_pct = (
            round(float(margen_bruto) / float(ingresos_netos) * 100, 2)
            if ingresos_netos > 0 else 0.0
        )

        total_gastos   = gastos_qs.aggregate(t=Sum("monto"))["t"] or Decimal("0")
        gastos_detalle = [
            {"categoria": g["categoria"], "total": float(g["total"]), "cantidad": g["cantidad"]}
            for g in gastos_qs.values("categoria")
            .annotate(total=Sum("monto"), cantidad=Count("id")).order_by("-total")
        ]

        _v = lambda qs: qs.aggregate(
            v=Sum(F("cantidad") * F("producto__precio_compra"), output_field=_dec)
        )["v"] or Decimal("0")
        perdidas_brutas      = _v(mov_qs.filter(tipo="dano"))
        valor_recuperado     = _v(mov_qs.filter(referencia_tipo="recuperacion_averia"))
        perdida_neta_averias = max(Decimal("0"), perdidas_brutas - valor_recuperado)

        utilidad_operativa     = margen_bruto - total_gastos - perdida_neta_averias
        utilidad_operativa_pct = (
            round(float(utilidad_operativa) / float(ingresos_netos) * 100, 2)
            if ingresos_netos > 0 else 0.0
        )

        return {
            "periodo": {"desde": str(fecha_ini), "hasta": str(fecha_fin), "tienda_id": tienda_id},
            "ingresos": {
                "ventas_brutas": float(ventas_brutas), "menos_descuentos": float(total_descuentos),
                "menos_devoluciones": float(total_devoluciones), "ingresos_netos": float(ingresos_netos),
                "impuestos_cobrados": float(total_impuestos),
                "num_ventas": num_ventas, "num_devoluciones": num_devoluciones,
            },
            "costo_ventas": float(cogs),
            "margen_bruto": float(margen_bruto), "margen_bruto_pct": margen_bruto_pct,
            "gastos_operativos": {"total": float(total_gastos), "detalle": gastos_detalle},
            "averias": {
                "perdidas_brutas": float(perdidas_brutas), "valor_recuperado": float(valor_recuperado),
                "perdida_neta": float(perdida_neta_averias),
            },
            "utilidad_operativa": float(utilidad_operativa),
            "utilidad_operativa_pct": utilidad_operativa_pct,
        }

    # ── Lógica reutilizada de FlujoCajaView ──────────────────
    def _get_flujo_caja(self, request, fecha_ini, fecha_fin, tienda_id):
        sesiones_qs = SesionCaja.objects.filter(
            fecha_apertura__date__gte=fecha_ini,
            fecha_apertura__date__lte=fecha_fin,
        ).select_related("tienda", "empleado").order_by("fecha_apertura")
        sesiones_qs = scope_qs(
            request, sesiones_qs, campo_empresa="tienda__empresa", tienda_id=tienda_id,
        )
        sesion_ids = list(sesiones_qs.values_list("id", flat=True))

        ventas_raw = (
            Venta.objects.filter(sesion_caja_id__in=sesion_ids, estado="completada")
            .values("sesion_caja_id", "metodo_pago").annotate(total=Sum("total"))
        )
        ventas_map = {}
        for v in ventas_raw:
            ventas_map.setdefault(v["sesion_caja_id"], {})[v["metodo_pago"]] = float(v["total"] or 0)

        gastos_map = {
            g["sesion_caja_id"]: float(g["total"] or 0)
            for g in Gasto.objects.filter(sesion_caja_id__in=sesion_ids)
            .values("sesion_caja_id").annotate(total=Sum("monto"))
        }
        abonos_map = {
            a["sesion_id"]: float(a["total"] or 0)
            for a in MovimientoCaja.objects.filter(sesion_id__in=sesion_ids, tipo="abono_separado")
            .values("sesion_id").annotate(total=Sum("monto"))
        }
        dev_map = {
            d["venta__sesion_caja_id"]: float(d["total"] or 0)
            for d in Devolucion.objects.filter(
                venta__sesion_caja_id__in=sesion_ids, estado="procesada"
            ).values("venta__sesion_caja_id").annotate(total=Sum("total_devuelto"))
        }

        sesiones = []
        for s in sesiones_qs:
            vm  = ventas_map.get(s.id, {})
            vef = vm.get("efectivo", 0); vtar = vm.get("tarjeta", 0)
            vtr = vm.get("transferencia", 0); vmx = vm.get("mixto", 0)
            ab  = abonos_map.get(s.id, 0)
            gas = gastos_map.get(s.id, 0)
            dev = dev_map.get(s.id, 0)
            ent = round(vef + vtar + vtr + vmx + ab, 2)
            sal = round(gas + dev, 2)
            sesiones.append({
                "sesion_id": s.id, "fecha": s.fecha_apertura.date().isoformat(),
                "tienda_id": s.tienda_id, "tienda_nombre": s.tienda.nombre,
                "empleado": f"{s.empleado.nombre} {s.empleado.apellido}" if s.empleado else None,
                "monto_inicial": float(s.monto_inicial),
                "entradas": {
                    "ventas_efectivo": vef, "ventas_tarjeta": vtar,
                    "ventas_transferencia": vtr, "ventas_mixto": vmx,
                    "abonos": ab, "total": ent,
                },
                "salidas": {"gastos": gas, "devoluciones": dev, "total": sal},
                "flujo_sesion": round(ent - sal, 2),
                "monto_final_sistema": float(s.monto_final_sistema) if s.monto_final_sistema is not None else None,
                "monto_final_real":    float(s.monto_final_real)    if s.monto_final_real    is not None else None,
                "diferencia":          float(s.diferencia)          if s.diferencia          is not None else None,
                "estado": s.estado,
            })

        total_e = round(sum(s["entradas"]["total"] for s in sesiones), 2)
        total_s = round(sum(s["salidas"]["total"]  for s in sesiones), 2)
        total_d = round(sum(s["diferencia"] or 0   for s in sesiones), 2)
        return {
            "periodo": {"desde": str(fecha_ini), "hasta": str(fecha_fin), "tienda_id": tienda_id},
            "resumen": {
                "total_entradas": total_e, "total_salidas": total_s,
                "flujo_neto": round(total_e - total_s, 2),
                "total_diferencias": total_d, "num_sesiones": len(sesiones),
            },
            "sesiones": sesiones,
        }
