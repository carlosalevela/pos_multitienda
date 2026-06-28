from decimal import Decimal

from django.db import transaction
from django.db.models import F, Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import generics, status
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from configuracion.models import ConfigTienda
from core.permissions import EsAdmin, EsAdminOSupervisor, EsAdminSupervisorOCajero, es_superadmin, get_empresa, scope_qs
from empresas.models import Empresa
from tiendas.models import Tienda
from ventas.models import DetalleVenta
from .models import (
    Categoria, Producto, Inventario,
    MovimientoInventario, generar_codigo_barras_interno,
)
from .serializers import (
    CategoriaSerializer, ProductoSerializer, ProductoSimpleSerializer,
    InventarioSerializer, AjusteInventarioSerializer,
    MovimientoInventarioSerializer, ImportarProductoItemSerializer,
)



def _mayoreo_habilitado(empresa) -> bool:
    """True si empresa o alguna de sus tiendas tiene mayoreo activo."""
    if not empresa:
        return False
    if empresa.maneja_mayoreo:
        return True
    return ConfigTienda.objects.filter(
        tienda__empresa=empresa, habilitar_mayoreo=True
    ).exists()


def _resolver_categoria(raw_nombre: str, empresa):
    nombre = ' '.join((raw_nombre or '').split())
    if not nombre:
        return None
    categoria = Categoria.objects.filter(
        nombre__iexact=nombre, empresa=empresa
    ).first()
    if not categoria:
        categoria = Categoria.objects.create(nombre=nombre, empresa=empresa)
    return categoria


# ── Categorías ────────────────────────────────────────────
class CategoriaListCreateView(generics.ListCreateAPIView):
    serializer_class = CategoriaSerializer

    def get_permissions(self):
        if self.request.method == "GET":
            return [IsAuthenticated()]
        return [EsAdminOSupervisor()]

    def get_queryset(self):
        if es_superadmin(self.request):
            empresa_id = self.request.query_params.get("empresa")
            qs = Categoria.objects.all()
            if empresa_id:
                qs = qs.filter(empresa_id=empresa_id)
            return qs.order_by("nombre")
        return Categoria.objects.filter(
            empresa=get_empresa(self.request)
        ).order_by("nombre")

    def perform_create(self, serializer):
        if es_superadmin(self.request):
            if not self.request.data.get("empresa"):
                raise PermissionDenied("El superadmin debe especificar una empresa.")
            serializer.save()
        else:
            serializer.save(empresa=get_empresa(self.request))


class CategoriaDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = CategoriaSerializer
    permission_classes = [EsAdmin]

    def get_queryset(self):
        if es_superadmin(self.request):
            return Categoria.objects.all()
        return Categoria.objects.filter(empresa=get_empresa(self.request))


# ── Productos ─────────────────────────────────────────────
class ProductoListCreateView(generics.ListCreateAPIView):
    serializer_class = ProductoSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [IsAuthenticated()]
        return [EsAdminOSupervisor()]

    def get_queryset(self):
        if es_superadmin(self.request):
            empresa_id = self.request.query_params.get("empresa")
            qs = Producto.objects.select_related("categoria", "empresa")
            if empresa_id:
                qs = qs.filter(empresa_id=empresa_id)
        else:
            qs = Producto.objects.select_related(
                "categoria", "empresa"
            ).filter(empresa=get_empresa(self.request))

        activo = self.request.query_params.get("activo", "true")
        if activo == "true":
            qs = qs.filter(activo=True)
        elif activo == "false":
            qs = qs.filter(activo=False)

        categoria = self.request.query_params.get("categoria")
        buscar    = self.request.query_params.get("q")
        tienda_id = self.request.query_params.get("tienda_id")

        if categoria: qs = qs.filter(categoria_id=categoria)
        if buscar:
            qs = qs.filter(
                Q(nombre__icontains=buscar) |
                Q(codigo_barras__icontains=buscar)
            )
        if tienda_id:
            empresa_filter = {} if es_superadmin(self.request) else \
                {"inventarios__tienda__empresa": get_empresa(self.request)}
            qs = qs.filter(
                inventarios__tienda_id=tienda_id,
                **empresa_filter,
            ).distinct()
        return qs

    @transaction.atomic
    def perform_create(self, serializer):
        if es_superadmin(self.request):
            empresa_id = self.request.data.get("empresa")
            if not empresa_id:
                raise PermissionDenied("El superadmin debe especificar una empresa.")
            empresa = Empresa.objects.get(id=empresa_id)
        else:
            empresa = get_empresa(self.request)

        categoria = _resolver_categoria(
            self.request.data.get('categoria_nombre', ''), empresa)

        extra = {}
        if _mayoreo_habilitado(empresa):
            precio_mayoreo = self.request.data.get('precio_mayoreo')
            if precio_mayoreo is not None:
                extra['precio_mayoreo'] = Decimal(str(precio_mayoreo))

        producto = serializer.save(
            empresa=empresa, categoria=categoria, **extra)

        tienda_id    = self.request.data.get('tienda_id')
        stock_actual = self.request.data.get('stock_actual', 0)
        stock_minimo = self.request.data.get('stock_minimo', 0)
        stock_maximo = self.request.data.get('stock_maximo', 0)

        if tienda_id:
            if not Tienda.objects.filter(id=tienda_id, empresa=empresa).exists():
                raise ValidationError(
                    {"tienda_id": "La tienda no pertenece a esta empresa."})
            Inventario.objects.update_or_create(
                producto=producto, tienda_id=tienda_id,
                defaults={
                    'stock_actual': Decimal(str(stock_actual)),
                    'stock_minimo': Decimal(str(stock_minimo)),
                    'stock_maximo': Decimal(str(stock_maximo)),
                }
            )


class ProductoDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = ProductoSerializer
    permission_classes = [EsAdminOSupervisor]

    def get_queryset(self):
        if es_superadmin(self.request):
            return Producto.objects.select_related("categoria", "empresa")
        return Producto.objects.select_related(
            "categoria", "empresa"
        ).filter(empresa=get_empresa(self.request))

    @transaction.atomic
    def partial_update(self, request, *args, **kwargs):
        producto = self.get_object()
        extra    = {}

        empresa = producto.empresa if es_superadmin(request) \
            else get_empresa(request)

        raw_cat = request.data.get('categoria_nombre', '')
        if raw_cat:
            extra['categoria'] = _resolver_categoria(raw_cat, empresa)

        if _mayoreo_habilitado(empresa):
            precio_mayoreo = request.data.get('precio_mayoreo')
            if precio_mayoreo is not None:
                extra['precio_mayoreo'] = Decimal(str(precio_mayoreo))
        else:
            extra['precio_mayoreo'] = None

        serializer = self.get_serializer(
            producto, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        producto = serializer.save(**extra)

        tienda_id    = request.data.get('tienda_id')
        stock_actual = request.data.get('stock_actual')
        stock_minimo = request.data.get('stock_minimo')
        stock_maximo = request.data.get('stock_maximo')

        if tienda_id:
            if not Tienda.objects.filter(id=tienda_id, empresa=empresa).exists():
                return Response(
                    {"tienda_id": "La tienda no pertenece a esta empresa."},
                    status=status.HTTP_403_FORBIDDEN)

            defaults = {}
            if stock_actual is not None:
                defaults['stock_actual'] = Decimal(str(stock_actual))
            if stock_minimo is not None:
                defaults['stock_minimo'] = Decimal(str(stock_minimo))
            if stock_maximo is not None:
                defaults['stock_maximo'] = Decimal(str(stock_maximo))
            if defaults:
                Inventario.objects.update_or_create(
                    producto=producto, tienda_id=tienda_id, defaults=defaults)

        return Response(
            self.get_serializer(producto).data,
            status=status.HTTP_200_OK)

    def destroy(self, request, *args, **kwargs):
        producto = self.get_object()
        producto.activo = False
        producto.save(update_fields=["activo"])
        return Response(
            {"detail": f"Producto '{producto.nombre}' desactivado."},
            status=status.HTTP_200_OK)


# ── Reactivar producto ────────────────────────────────────
class ReactivarProductoView(APIView):
    permission_classes = [EsAdmin]

    def patch(self, request, pk):
        try:
            producto = Producto.objects.get(pk=pk) \
                if es_superadmin(request) \
                else Producto.objects.get(
                    pk=pk, empresa=get_empresa(request))
        except Producto.DoesNotExist:
            return Response(
                {"error": "Producto no encontrado."},
                status=status.HTTP_404_NOT_FOUND)

        if producto.activo:
            return Response(
                {"error": "El producto ya está activo."},
                status=status.HTTP_400_BAD_REQUEST)

        producto.activo = True
        producto.save(update_fields=["activo"])
        return Response({
            "detail": f"Producto '{producto.nombre}' reactivado.",
            "id":     producto.id,
            "nombre": producto.nombre,
            "activo": producto.activo,
        })


# ── Búsqueda POS ──────────────────────────────────────────
class BuscarProductoPOSView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        q         = request.query_params.get("q", "").strip()
        tienda_id = request.query_params.get("tienda_id")

        if not q:
            return Response(
                {"error": "Ingresa un término de búsqueda."}, status=400)

        if es_superadmin(request):
            productos = Producto.objects.select_related("empresa").filter(
                Q(codigo_barras=q) | Q(nombre__icontains=q),
                activo=True,
            )[:10]
        else:
            productos = Producto.objects.select_related("empresa").filter(
                Q(codigo_barras=q) | Q(nombre__icontains=q),
                activo=True,
                empresa=get_empresa(request),
            )[:10]

        data = []
        for p in productos:
            item = ProductoSimpleSerializer(
                p, context={"request": request}).data
            if tienda_id:
                filtro = {"producto": p, "tienda_id": tienda_id}
                if not es_superadmin(request):
                    filtro["tienda__empresa"] = get_empresa(request)
                inv = Inventario.objects.filter(**filtro).first()
                item["stock_actual"] = float(inv.stock_actual) if inv else 0
                item["alerta_stock"] = (
                    "agotado" if not inv or inv.stock_actual <= 0
                    else "bajo" if inv.stock_actual <= inv.stock_minimo
                    else "ok"
                )
            data.append(item)
        return Response(data)


# ── Inventario ────────────────────────────────────────────
class InventarioListView(generics.ListAPIView):
    serializer_class   = InventarioSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = Inventario.objects.select_related(
            "producto", "producto__categoria", "tienda")
        qs = scope_qs(self.request, qs)

        tienda_id  = self.request.query_params.get("tienda_id")
        alerta     = self.request.query_params.get("alerta")
        activo     = self.request.query_params.get("activo", "true")
        categoria  = self.request.query_params.get("categoria")

        if activo == "true":    qs = qs.filter(producto__activo=True)
        elif activo == "false": qs = qs.filter(producto__activo=False)
        if tienda_id:  qs = qs.filter(tienda_id=tienda_id)
        if categoria:  qs = qs.filter(producto__categoria_id=categoria)
        if alerta == "bajo":
            qs = qs.filter(
                stock_actual__lte=F("stock_minimo"),
                stock_actual__gt=0)
        elif alerta == "agotado":
            qs = qs.filter(stock_actual__lte=0)

        return qs


# ── Ajuste de inventario ──────────────────────────────────
class AjustarInventarioView(APIView):
    permission_classes = [EsAdminOSupervisor]

    @transaction.atomic
    def post(self, request, producto_id, tienda_id):
        serializer = AjusteInventarioSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        try:
            filtro = {"producto_id": producto_id, "tienda_id": tienda_id}
            if not es_superadmin(request):
                filtro["tienda__empresa"] = get_empresa(request)
            inv = Inventario.objects.select_for_update().get(**filtro)
        except Inventario.DoesNotExist:
            return Response(
                {"error": "Producto no existe en esta tienda."}, status=404)

        tipo        = serializer.validated_data["tipo"]
        cantidad    = serializer.validated_data["cantidad"]
        observacion = serializer.validated_data.get("observacion", "")

        if tipo == "entrada":
            inv.stock_actual += cantidad
        elif tipo == "salida":
            if inv.stock_actual < cantidad:
                return Response(
                    {"error": "Stock insuficiente."}, status=400)
            inv.stock_actual -= cantidad
        elif tipo == "ajuste":
            inv.stock_actual = cantidad

        inv.save()

        MovimientoInventario.objects.create(
            producto_id=producto_id, tienda_id=tienda_id,
            empleado=request.user, tipo=tipo, cantidad=cantidad,
            referencia_tipo="manual", observacion=observacion,
        )

        return Response({
            "detail":       "Stock actualizado correctamente.",
            "stock_actual": float(inv.stock_actual),
            "tipo":         tipo,
        })


# ── Movimientos de producto ───────────────────────────────
class MovimientosProductoView(generics.ListAPIView):
    serializer_class   = MovimientoInventarioSerializer
    permission_classes = [EsAdminOSupervisor]

    def get_queryset(self):
        filtro = {
            "producto_id": self.kwargs["producto_id"],
            "tienda_id":   self.kwargs["tienda_id"],
        }
        if not es_superadmin(self.request):
            filtro["tienda__empresa"] = get_empresa(self.request)
        return MovimientoInventario.objects.filter(
            **filtro
        ).select_related("producto", "empleado").order_by("-created_at")


# ── Top productos ─────────────────────────────────────────
class TopProductosView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        tienda_id = request.query_params.get('tienda_id')
        fecha_ini = request.query_params.get('fecha_ini')
        fecha_fin = request.query_params.get('fecha_fin')
        limite    = int(request.query_params.get('limite', 20))

        qs = scope_qs(
            request,
            DetalleVenta.objects.filter(venta__estado='completada'),
            campo_empresa="venta__tienda__empresa",
        )

        if request.user.rol == 'cajero':
            qs = qs.filter(venta__tienda_id=request.user.tienda_id)
        elif tienda_id:
            qs = qs.filter(venta__tienda_id=tienda_id)

        if fecha_ini: qs = qs.filter(venta__created_at__date__gte=fecha_ini)
        if fecha_fin: qs = qs.filter(venta__created_at__date__lte=fecha_fin)

        resultado = (
            qs.values('producto__nombre', 'producto__categoria__nombre')
            .annotate(
                total_vendido  = Sum('cantidad'),
                total_ingresos = Sum('subtotal'),
            )
            .order_by('-total_ingresos')[:limite]
        )

        return Response([
            {
                'producto':       r['producto__nombre']            or 'Sin nombre',
                'categoria':      r['producto__categoria__nombre'] or 'Sin categoría',
                'total_vendido':  float(r['total_vendido']  or 0),
                'total_ingresos': float(r['total_ingresos'] or 0),
            }
            for r in resultado
        ])


# ── Importar productos desde Excel (batch) ────────────────
class ImportarProductosView(APIView):
    permission_classes = [EsAdminOSupervisor]

    _COLUMN_MAP = {
        "nombre":           "nombre",
        "descripcion":      "descripcion",
        "descripción":      "descripcion",
        "codigo_barras":    "codigo_barras",
        "código de barras": "codigo_barras",
        "codigo barras":    "codigo_barras",
        "codigo":           "codigo_barras",
        "código":           "codigo_barras",
        "categoria":        "categoria_nombre",
        "categoría":        "categoria_nombre",
        "categoria_nombre": "categoria_nombre",
        "precio_venta":     "precio_venta",
        "precio venta":     "precio_venta",
        "precio de venta":  "precio_venta",
        "precio_compra":    "precio_compra",
        "precio compra":    "precio_compra",
        "precio de compra": "precio_compra",
        "precio_mayoreo":   "precio_mayoreo",
        "precio mayoreo":   "precio_mayoreo",
        "stock_actual":     "stock_actual",
        "stock actual":     "stock_actual",
        "stock":            "stock_actual",
        "existencia":       "stock_actual",
        "stock_minimo":     "stock_minimo",
        "stock mínimo":     "stock_minimo",
        "stock minimo":     "stock_minimo",
        "stock_maximo":     "stock_maximo",
        "stock máximo":     "stock_maximo",
        "stock maximo":     "stock_maximo",
    }

    def _parsear_excel(self, archivo):
        import openpyxl
        try:
            wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            ws = wb.active
            filas = list(ws.iter_rows(values_only=True))
        except Exception as e:
            raise ValueError(f"No se pudo leer el archivo Excel: {e}")

        if not filas:
            raise ValueError("El archivo Excel está vacío.")

        raw_headers = [
            str(h).strip().lower() if h is not None else ""
            for h in filas[0]
        ]

        field_indices = {}
        for idx, h in enumerate(raw_headers):
            campo = self._COLUMN_MAP.get(h)
            if campo and campo not in field_indices:
                field_indices[campo] = idx

        if "nombre" not in field_indices:
            cols = ", ".join(h for h in raw_headers if h)
            raise ValueError(
                f"Columna 'nombre' no encontrada. "
                f"Columnas detectadas: {cols or '(ninguna)'}"
            )

        productos = []
        for fila in filas[1:]:
            if all(v is None or str(v).strip() == "" for v in fila):
                continue
            item = {}
            for campo, idx in field_indices.items():
                valor = fila[idx] if idx < len(fila) else None
                if valor is not None and str(valor).strip() != "":
                    item[campo] = valor
            productos.append(item)

        return productos

    def post(self, request):
        if es_superadmin(request):
            empresa_id = request.data.get("empresa")
            if not empresa_id:
                return Response(
                    {"detail": "El superadmin debe especificar una empresa."},
                    status=status.HTTP_400_BAD_REQUEST)
            try:
                empresa = Empresa.objects.get(id=empresa_id)
            except Empresa.DoesNotExist:
                return Response(
                    {"detail": "Empresa no encontrada."},
                    status=status.HTTP_404_NOT_FOUND)
        else:
            empresa = get_empresa(request)

        tienda_id = request.data.get("tienda_id")
        tienda    = None
        if tienda_id:
            try:
                tienda = Tienda.objects.get(id=tienda_id, empresa=empresa)
            except Tienda.DoesNotExist:
                return Response(
                    {"detail": "La tienda no pertenece a esta empresa."},
                    status=status.HTTP_400_BAD_REQUEST)

        archivo = request.FILES.get("archivo")
        if archivo:
            nombre_archivo = archivo.name.lower()
            if not (nombre_archivo.endswith(".xlsx") or nombre_archivo.endswith(".xls")):
                return Response(
                    {"detail": "Solo se aceptan archivos .xlsx o .xls."},
                    status=status.HTTP_400_BAD_REQUEST)
            try:
                productos_data = self._parsear_excel(archivo)
            except ValueError as e:
                return Response(
                    {"detail": str(e)},
                    status=status.HTTP_400_BAD_REQUEST)
        else:
            productos_data = request.data.get("productos", [])

        if not isinstance(productos_data, list) or len(productos_data) == 0:
            return Response(
                {"detail": "Sin datos: envía un archivo Excel (.xlsx) "
                           "o el campo 'productos' como lista no vacía."},
                status=status.HTTP_400_BAD_REQUEST)

        creados    = 0
        fallidos   = 0
        resultados = []

        for i, item in enumerate(productos_data):
            fila = i + 2

            ser = ImportarProductoItemSerializer(data=item)
            if not ser.is_valid():
                fallidos += 1
                resultados.append({
                    "fila":    fila,
                    "nombre":  item.get("nombre", "(vacío)"),
                    "success": False,
                    "error":   str(ser.errors),
                })
                continue

            d = ser.validated_data
            try:
                with transaction.atomic():
                    categoria = _resolver_categoria(
                        d.get("categoria_nombre", ""), empresa)

                    codigo = (d.get("codigo_barras") or "").strip() or None
                    if not codigo:
                        codigo = generar_codigo_barras_interno()

                    precio_mayoreo = (
                        d.get("precio_mayoreo") if empresa.maneja_mayoreo else None
                    )

                    producto = Producto.objects.create(
                        empresa         = empresa,
                        categoria       = categoria,
                        nombre          = d["nombre"],
                        descripcion     = d.get("descripcion", ""),
                        codigo_barras   = codigo,
                        precio_compra   = d.get("precio_compra", 0),
                        precio_venta    = d.get("precio_venta",  0),
                        precio_mayoreo  = precio_mayoreo,
                        unidad_medida   = "unidad",
                        aplica_impuesto = False,
                    )

                    if tienda:
                        inv_defaults = {
                            "stock_actual": d.get("stock_actual", 0),
                            "stock_minimo": d.get("stock_minimo", 0),
                        }
                        if d.get("stock_maximo") is not None:
                            inv_defaults["stock_maximo"] = d["stock_maximo"]
                        Inventario.objects.update_or_create(
                            producto = producto,
                            tienda   = tienda,
                            defaults = inv_defaults,
                        )

                creados += 1
                resultados.append({
                    "fila":    fila,
                    "nombre":  d["nombre"],
                    "success": True,
                    "id":      producto.id,
                })

            except Exception as e:
                fallidos += 1
                msg = str(e)
                if "unique" in msg.lower():
                    msg = f"Código de barras '{codigo}' ya existe."
                resultados.append({
                    "fila":    fila,
                    "nombre":  d["nombre"],
                    "success": False,
                    "error":   msg,
                })

        return Response({
            "creados":    creados,
            "fallidos":   fallidos,
            "total":      len(productos_data),
            "resultados": resultados,
        }, status=status.HTTP_207_MULTI_STATUS)


# ── Dashboard KPIs de inventario ──────────────────────────
class DashboardInventarioView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        inicio_mes = timezone.now().replace(
            day=1, hour=0, minute=0, second=0, microsecond=0)

        tienda_id = request.query_params.get("tienda_id")

        inv_qs, mov_qs = scope_qs(
            request,
            Inventario.objects.select_related("producto", "tienda"),
            MovimientoInventario.objects.all(),
            tienda_id=tienda_id,
        )
        inv_qs = inv_qs.filter(producto__activo=True)

        total_valor = inv_qs.aggregate(
            valor=Sum(F("stock_actual") * F("producto__precio_venta"))
        )["valor"] or 0

        alertas_bajo = inv_qs.filter(
            stock_actual__lte=F("stock_minimo"),
            stock_actual__gt=0,
        ).count()

        alertas_criticas = inv_qs.filter(stock_actual__lte=0).count()

        mov_mes      = mov_qs.filter(created_at__gte=inicio_mes)
        ingresos_mes = mov_mes.filter(
            tipo="entrada"
        ).exclude(
            referencia_tipo="recuperacion_averia"
        ).aggregate(total=Sum("cantidad"))["total"] or 0

        perdidas_brutas = mov_mes.filter(
            tipo="dano"
        ).select_related("producto").aggregate(
            valor=Sum(F("cantidad") * F("producto__precio_compra"))
        )["valor"] or 0

        recuperado = mov_mes.filter(
            referencia_tipo="recuperacion_averia"
        ).select_related("producto").aggregate(
            valor=Sum(F("cantidad") * F("producto__precio_compra"))
        )["valor"] or 0

        perdidas_valor = max(0.0, float(perdidas_brutas) - float(recuperado))

        prod_qs      = scope_qs(request, Producto.objects.all(), campo_empresa="empresa")
        productos_mes = prod_qs.filter(
            created_at__gte=inicio_mes, activo=True
        ).count()

        return Response({
            "total_valor_inventario":    float(total_valor),
            "alertas_stock_bajo":        alertas_bajo,
            "alertas_criticas":          alertas_criticas,
            "ingresos_mes_unidades":     float(ingresos_mes),
            "productos_registrados_mes": productos_mes,
            "perdidas_mes_valor":        float(perdidas_valor),
        })


# ── Feed global de movimientos recientes ─────────────────
class MovimientosRecientesView(generics.ListAPIView):
    serializer_class   = MovimientoInventarioSerializer
    permission_classes = [EsAdminSupervisorOCajero]

    def get_queryset(self):
        qs = MovimientoInventario.objects.select_related(
            "producto", "tienda", "empleado"
        )
        qs = scope_qs(self.request, qs)

        tienda_id = self.request.query_params.get("tienda_id")
        tipo      = self.request.query_params.get("tipo")
        limite    = int(self.request.query_params.get("limite", 20))

        # Cajero solo ve los movimientos de su propia tienda
        if self.request.user.rol == "cajero":
            qs = qs.filter(tienda_id=self.request.user.tienda_id)
        elif tienda_id:
            qs = qs.filter(tienda_id=tienda_id)

        if tipo: qs = qs.filter(tipo=tipo)

        return qs.order_by("-created_at")[:limite]


# ── Averías ───────────────────────────────────────────────
class AveriasView(generics.ListAPIView):
    serializer_class   = InventarioSerializer
    permission_classes = [EsAdminOSupervisor]

    def get_queryset(self):
        qs = Inventario.objects.select_related(
            "producto", "producto__categoria", "tienda"
        ).filter(stock_averias__gt=0, producto__activo=True)
        qs = scope_qs(self.request, qs)

        tienda_id = self.request.query_params.get("tienda_id")
        if tienda_id:
            qs = qs.filter(tienda_id=tienda_id)
        return qs.order_by("-stock_averias")


class RecuperarAveriaView(APIView):
    permission_classes = [EsAdminOSupervisor]

    @transaction.atomic
    def post(self, request, producto_id, tienda_id):
        accion   = request.data.get("accion")
        cantidad = request.data.get("cantidad")

        if accion not in ("recuperar", "descartar"):
            return Response(
                {"error": "accion debe ser 'recuperar' o 'descartar'."},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            cantidad = Decimal(str(cantidad))
            if cantidad <= 0:
                raise ValueError()
        except (TypeError, ValueError):
            return Response(
                {"error": "Cantidad inválida."}, status=status.HTTP_400_BAD_REQUEST)

        filtro = {"producto_id": producto_id, "tienda_id": tienda_id}
        if not es_superadmin(request):
            filtro["tienda__empresa"] = get_empresa(request)

        try:
            inv = Inventario.objects.select_for_update().get(**filtro)
        except Inventario.DoesNotExist:
            return Response(
                {"error": "Producto no encontrado en esta tienda."},
                status=status.HTTP_404_NOT_FOUND)

        if cantidad > inv.stock_averias:
            return Response(
                {"error": f"Stock en averías insuficiente. "
                          f"Disponible: {inv.stock_averias}."},
                status=status.HTTP_400_BAD_REQUEST)

        inv.stock_averias -= cantidad
        if accion == "recuperar":
            inv.stock_actual += cantidad
        inv.save(update_fields=["stock_averias", "stock_actual"])

        MovimientoInventario.objects.create(
            producto_id=producto_id, tienda_id=tienda_id,
            empleado=request.user,
            tipo="entrada" if accion == "recuperar" else "salida",
            cantidad=cantidad,
            referencia_tipo=(
                "recuperacion_averia" if accion == "recuperar"
                else "descarte_averia"),
            observacion=(
                f"Recuperado de averías → stock normal"
                if accion == "recuperar"
                else "Descartado de averías (baja definitiva)"),
        )

        return Response({
            "detail": (
                "Unidades recuperadas al stock normal."
                if accion == "recuperar"
                else "Unidades dadas de baja definitivamente."),
            "stock_actual":  float(inv.stock_actual),
            "stock_averias": float(inv.stock_averias),
        })


# ── Exportar inventario a Excel ───────────────────────────
class ExportarInventarioView(APIView):
    permission_classes = [EsAdminOSupervisor]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        qs = Inventario.objects.select_related(
            "producto", "producto__categoria", "tienda", "tienda__empresa"
        )
        qs = scope_qs(request, qs)

        tienda_id = request.query_params.get("tienda_id")
        alerta    = request.query_params.get("alerta")
        activo    = request.query_params.get("activo", "true")

        if activo == "true":
            qs = qs.filter(producto__activo=True)
        elif activo == "false":
            qs = qs.filter(producto__activo=False)
        if tienda_id:
            qs = qs.filter(tienda_id=tienda_id)
        if alerta == "bajo":
            qs = qs.filter(
                stock_actual__lte=F("stock_minimo"), stock_actual__gt=0)
        elif alerta == "agotado":
            qs = qs.filter(stock_actual__lte=0)

        inventarios = list(qs.order_by("tienda__nombre", "producto__nombre"))

        if not es_superadmin(request):
            empresa_nombre = get_empresa(request).nombre
        elif inventarios:
            empresa_nombre = inventarios[0].tienda.empresa.nombre \
                if inventarios[0].tienda.empresa else "Todas las empresas"
        else:
            empresa_nombre = "Todas las empresas"

        if tienda_id and inventarios:
            tienda_label = inventarios[0].tienda.nombre
        elif tienda_id:
            tienda_label = f"Tienda #{tienda_id}"
        else:
            tienda_label = "Todas las tiendas"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"

        C_AZUL_OSCURO = "1E3A5F"
        C_AZUL_MEDIO  = "2E86AB"
        C_FILA_PAR    = "EBF5FB"
        C_AGOTADO     = "FADBD8"
        C_BAJO        = "FEF9E7"
        C_TOTAL       = "D5D8DC"
        C_OK          = "D5F5E3"

        def fill(color):
            return PatternFill("solid", fgColor=color)

        def border():
            thin = Side(style="thin", color="BFBFBF")
            return Border(left=thin, right=thin, top=thin, bottom=thin)

        f_titulo   = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
        f_sub      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        f_header   = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        f_normal   = Font(name="Calibri", size=10)
        f_bold     = Font(name="Calibri", bold=True, size=10)
        f_rojo     = Font(name="Calibri", bold=True, size=10, color="C0392B")
        f_amarillo = Font(name="Calibri", bold=True, size=10, color="9A7D0A")
        f_verde    = Font(name="Calibri", bold=True, size=10, color="1E8449")

        ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
        al = Alignment(horizontal="left",   vertical="center")
        ar = Alignment(horizontal="right",  vertical="center")

        COLUMNAS = [
            ("#",                     5,  "center"),
            ("Código de Barras",     18,  "left"),
            ("Nombre del Producto",  35,  "left"),
            ("Categoría",            20,  "left"),
            ("Precio Compra",        15,  "right"),
            ("Precio Venta",         14,  "right"),
            ("Precio Mayoreo",       14,  "right"),
            ("Stock Actual",         13,  "right"),
            ("Stock Mínimo",         13,  "right"),
            ("Stock Máximo",         13,  "right"),
            ("Valor Inventario",     18,  "right"),
            ("Estado",               12,  "center"),
            ("Tienda",               20,  "left"),
            ("Última Actualización", 20,  "center"),
        ]
        N = len(COLUMNAS)
        ultima_col = get_column_letter(N)
        now = timezone.now()

        ws.row_dimensions[1].height = 32
        ws.merge_cells(f"A1:{ultima_col}1")
        c = ws["A1"]
        c.value     = f"REPORTE DE INVENTARIO  —  {empresa_nombre.upper()}"
        c.font      = f_titulo
        c.fill      = fill(C_AZUL_OSCURO)
        c.alignment = ac

        ws.row_dimensions[2].height = 20
        ws.merge_cells(f"A2:{ultima_col}2")
        c = ws["A2"]
        c.value     = (
            f"Generado: {now.strftime('%d/%m/%Y %H:%M')}  |  "
            f"{tienda_label}  |  "
            f"Registros: {len(inventarios)}"
        )
        c.font      = f_sub
        c.fill      = fill(C_AZUL_MEDIO)
        c.alignment = ac

        ws.row_dimensions[3].height = 22
        for i, (nombre, ancho, _) in enumerate(COLUMNAS, start=1):
            col_letra = get_column_letter(i)
            ws.column_dimensions[col_letra].width = ancho
            c = ws.cell(row=3, column=i, value=nombre)
            c.font      = f_header
            c.fill      = fill(C_AZUL_OSCURO)
            c.alignment = ac
            c.border    = border()

        total_valor = 0.0

        for num, inv in enumerate(inventarios, start=1):
            fila = num + 3
            ws.row_dimensions[fila].height = 16

            stock      = float(inv.stock_actual)
            stock_min  = float(inv.stock_minimo)
            stock_max  = float(inv.stock_maximo)
            p_compra   = float(inv.producto.precio_compra)
            p_venta    = float(inv.producto.precio_venta)
            p_mayoreo  = float(inv.producto.precio_mayoreo) \
                         if inv.producto.precio_mayoreo else None
            valor      = stock * p_venta
            total_valor += valor

            if stock <= 0:
                estado       = "Agotado"
                fill_fila    = fill(C_AGOTADO)
                font_estado  = f_rojo
            elif stock <= stock_min:
                estado       = "Stock Bajo"
                fill_fila    = fill(C_BAJO)
                font_estado  = f_amarillo
            else:
                estado       = "OK"
                fill_fila    = fill(C_FILA_PAR) if num % 2 == 0 else None
                font_estado  = f_verde

            categoria   = inv.producto.categoria.nombre \
                          if inv.producto.categoria else "Sin categoría"
            actualizado = inv.updated_at.strftime("%d/%m/%Y %H:%M") \
                          if inv.updated_at else ""

            celdas = [
                (num,                              "center", f_normal,   None),
                (inv.producto.codigo_barras or "", "left",   f_normal,   None),
                (inv.producto.nombre,              "left",   f_bold,     None),
                (categoria,                        "left",   f_normal,   None),
                (p_compra,                         "right",  f_normal,   "#,##0.00"),
                (p_venta,                          "right",  f_normal,   "#,##0.00"),
                (p_mayoreo if p_mayoreo else "",   "right",  f_normal,   "#,##0.00" if p_mayoreo else None),
                (stock,                            "right",  f_normal,   "#,##0.00"),
                (stock_min,                        "right",  f_normal,   "#,##0.00"),
                (stock_max,                        "right",  f_normal,   "#,##0.00"),
                (valor,                            "right",  f_bold,     "#,##0.00"),
                (estado,                           "center", font_estado, None),
                (inv.tienda.nombre,                "left",   f_normal,   None),
                (actualizado,                      "center", f_normal,   None),
            ]

            for col_idx, (valor_celda, alin, fuente, fmt) in enumerate(celdas, start=1):
                c = ws.cell(row=fila, column=col_idx, value=valor_celda)
                c.font   = fuente
                c.border = border()
                c.alignment = (
                    ac if alin == "center" else
                    ar if alin == "right"  else al
                )
                if fill_fila and col_idx != 12:
                    c.fill = fill_fila
                elif col_idx == 12:
                    if estado == "Agotado":
                        c.fill = fill(C_AGOTADO)
                    elif estado == "Stock Bajo":
                        c.fill = fill(C_BAJO)
                    elif fill_fila:
                        c.fill = fill_fila
                if fmt:
                    c.number_format = fmt

        fila_total = len(inventarios) + 4
        ws.row_dimensions[fila_total].height = 20
        ws.merge_cells(f"A{fila_total}:J{fila_total}")
        c = ws.cell(row=fila_total, column=1, value="VALOR TOTAL DEL INVENTARIO")
        c.font      = f_bold
        c.fill      = fill(C_TOTAL)
        c.alignment = ar
        c.border    = border()

        c = ws.cell(row=fila_total, column=11, value=total_valor)
        c.font          = f_bold
        c.fill          = fill(C_TOTAL)
        c.alignment     = ar
        c.number_format = "#,##0.00"
        c.border        = border()

        for col_idx in range(12, N + 1):
            c = ws.cell(row=fila_total, column=col_idx, value="")
            c.fill   = fill(C_TOTAL)
            c.border = border()

        ws.freeze_panes = "A4"

        # ── Hoja 2: Averías ───────────────────────────────────
        ws2 = wb.create_sheet(title="Averías")

        COL_AV = [
            ("#",                    5,  "center"),
            ("Código de Barras",    18,  "left"),
            ("Nombre del Producto", 35,  "left"),
            ("Categoría",           18,  "left"),
            ("Tienda",              20,  "left"),
            ("Unidades Dañadas",    16,  "right"),
            ("Precio Compra",       14,  "right"),
            ("Valor Pérdida",       16,  "right"),
        ]
        N2        = len(COL_AV)
        ultima2   = get_column_letter(N2)
        C_ROJO    = "B03A2E"
        C_ROJO_L  = "FADBD8"
        C_ROJO_M  = "E74C3C"

        ws2.row_dimensions[1].height = 32
        ws2.merge_cells(f"A1:{ultima2}1")
        c = ws2["A1"]
        c.value     = f"REPORTE DE AVERÍAS  —  {empresa_nombre.upper()}"
        c.font      = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=C_ROJO)
        c.alignment = ac

        ws2.row_dimensions[2].height = 20
        ws2.merge_cells(f"A2:{ultima2}2")
        c = ws2["A2"]
        c.value     = (
            f"Generado: {now.strftime('%d/%m/%Y %H:%M')}  |  "
            f"{tienda_label}  |  Solo productos con stock dañado"
        )
        c.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=C_ROJO_M)
        c.alignment = ac

        ws2.row_dimensions[3].height = 22
        for i, (nombre_col, ancho, _) in enumerate(COL_AV, start=1):
            col_letra = get_column_letter(i)
            ws2.column_dimensions[col_letra].width = ancho
            c = ws2.cell(row=3, column=i, value=nombre_col)
            c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor=C_ROJO)
            c.alignment = ac
            c.border    = border()

        averias = [inv for inv in inventarios if float(inv.stock_averias) > 0]
        total_perdida = 0.0

        for num, inv in enumerate(averias, start=1):
            fila2 = num + 3
            ws2.row_dimensions[fila2].height = 16
            unidades  = float(inv.stock_averias)
            p_compra  = float(inv.producto.precio_compra)
            perdida   = unidades * p_compra
            total_perdida += perdida
            categoria = inv.producto.categoria.nombre \
                        if inv.producto.categoria else "Sin categoría"
            fondo = PatternFill("solid", fgColor=C_ROJO_L) if num % 2 == 0 else None

            celdas2 = [
                (num,                              "center", f_normal,  None),
                (inv.producto.codigo_barras or "", "left",   f_normal,  None),
                (inv.producto.nombre,              "left",   f_bold,    None),
                (categoria,                        "left",   f_normal,  None),
                (inv.tienda.nombre,                "left",   f_normal,  None),
                (unidades,                         "right",  f_rojo,    "#,##0.00"),
                (p_compra,                         "right",  f_normal,  "#,##0.00"),
                (perdida,                          "right",  f_rojo,    "#,##0.00"),
            ]
            for col_idx, (val_c, alin, fuente, fmt) in enumerate(celdas2, start=1):
                c = ws2.cell(row=fila2, column=col_idx, value=val_c)
                c.font      = fuente
                c.border    = border()
                c.alignment = (
                    ac if alin == "center" else
                    ar if alin == "right"  else al
                )
                if fondo:
                    c.fill = fondo
                if fmt:
                    c.number_format = fmt

        if not averias:
            fila_vacia = 4
            ws2.merge_cells(f"A{fila_vacia}:{ultima2}{fila_vacia}")
            c = ws2.cell(row=fila_vacia, column=1, value="No hay productos con averías registradas.")
            c.font      = Font(name="Calibri", size=11, color="1E8449")
            c.alignment = ac
        else:
            fila_tot2 = len(averias) + 4
            ws2.row_dimensions[fila_tot2].height = 20
            ws2.merge_cells(f"A{fila_tot2}:G{fila_tot2}")
            c = ws2.cell(row=fila_tot2, column=1, value="VALOR TOTAL DE PÉRDIDAS POR AVERÍAS")
            c.font      = Font(name="Calibri", bold=True, size=10)
            c.fill      = PatternFill("solid", fgColor=C_TOTAL)
            c.alignment = ar
            c.border    = border()
            c = ws2.cell(row=fila_tot2, column=8, value=total_perdida)
            c.font          = Font(name="Calibri", bold=True, size=10, color=C_ROJO)
            c.fill          = PatternFill("solid", fgColor=C_TOTAL)
            c.alignment     = ar
            c.number_format = "#,##0.00"
            c.border        = border()

        ws2.freeze_panes = "A4"

        # ── Hoja 3: Historial de movimientos de daño ─────────
        mov_qs = MovimientoInventario.objects.filter(tipo="dano").select_related(
            "producto", "tienda", "empleado"
        )
        if not es_superadmin(request):
            mov_qs = mov_qs.filter(tienda__empresa=get_empresa(request))
        if tienda_id:
            mov_qs = mov_qs.filter(tienda_id=tienda_id)
        movimientos = list(mov_qs.order_by("-created_at")[:500])

        if movimientos:
            ws3 = wb.create_sheet(title="Historial Daños")

            COL_MOV = [
                ("#",              5,  "center"),
                ("Fecha",         18,  "center"),
                ("Producto",      32,  "left"),
                ("Tienda",        20,  "left"),
                ("Empleado",      22,  "left"),
                ("Unidades",      12,  "right"),
                ("Observación",   40,  "left"),
            ]
            N3      = len(COL_MOV)
            ultima3 = get_column_letter(N3)

            ws3.row_dimensions[1].height = 32
            ws3.merge_cells(f"A1:{ultima3}1")
            c = ws3["A1"]
            c.value     = f"HISTORIAL DE DAÑOS  —  {empresa_nombre.upper()}"
            c.font      = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor=C_ROJO)
            c.alignment = ac

            ws3.row_dimensions[2].height = 20
            ws3.merge_cells(f"A2:{ultima3}2")
            c = ws3["A2"]
            c.value     = (
                f"Generado: {now.strftime('%d/%m/%Y %H:%M')}  |  "
                f"{tienda_label}  |  Últimos {len(movimientos)} registros"
            )
            c.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor=C_ROJO_M)
            c.alignment = ac

            ws3.row_dimensions[3].height = 22
            for i, (nombre_col, ancho, _) in enumerate(COL_MOV, start=1):
                ws3.column_dimensions[get_column_letter(i)].width = ancho
                c = ws3.cell(row=3, column=i, value=nombre_col)
                c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
                c.fill      = PatternFill("solid", fgColor=C_ROJO)
                c.alignment = ac
                c.border    = border()

            for num, mov in enumerate(movimientos, start=1):
                fila3 = num + 3
                ws3.row_dimensions[fila3].height = 16
                empleado_nombre = (
                    f"{mov.empleado.nombre} {mov.empleado.apellido}"
                    if mov.empleado else "—"
                )
                fondo3 = PatternFill("solid", fgColor=C_ROJO_L) if num % 2 == 0 else None
                celdas3 = [
                    (num,                                "center", f_normal, None),
                    (mov.created_at.strftime("%d/%m/%Y %H:%M"), "center", f_normal, None),
                    (mov.producto.nombre,                "left",   f_bold,  None),
                    (mov.tienda.nombre,                  "left",   f_normal, None),
                    (empleado_nombre,                    "left",   f_normal, None),
                    (float(mov.cantidad),                "right",  f_rojo,   "#,##0.00"),
                    (mov.observacion or "—",             "left",   f_normal, None),
                ]
                for col_idx, (val_c, alin, fuente, fmt) in enumerate(celdas3, start=1):
                    c = ws3.cell(row=fila3, column=col_idx, value=val_c)
                    c.font      = fuente
                    c.border    = border()
                    c.alignment = (
                        ac if alin == "center" else
                        ar if alin == "right"  else al
                    )
                    if fondo3:
                        c.fill = fondo3
                    if fmt:
                        c.number_format = fmt

            ws3.freeze_panes = "A4"

        fecha_str = now.strftime("%Y%m%d_%H%M")
        response  = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            )
        )
        response["Content-Disposition"] = (
            f'attachment; filename="inventario_{fecha_str}.xlsx"'
        )
        wb.save(response)
        return response
