"""
Script de diagnóstico: genera el xlsx directamente a disco sin Django/HTTP.
Ejecutar con: .\venv\Scripts\python test_excel_gen.py
"""
import sys
import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

import django
django.setup()

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Mismos valores que en views.py ───────────────────────────
_NAVY    = "1E3A5F"
_BLUE    = "2E6DA4"
_GRAY    = "F2F2F2"
_GREEN   = "1E7E34"
_RED     = "B22222"
_FMT_CUR = '#,##0.00'
_FMT_PCT = '0.00%'


def _hdr(ws, row, col, value, bold=True, bg=None, color="FFFFFF", size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color, size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return cell


def _val(ws, row, col, value, bold=False, color="000000", fmt=_FMT_CUR, bg=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


# ── Datos de prueba ───────────────────────────────────────────
data = {
    "periodo": {"desde": "2026-06-01", "hasta": "2026-06-22", "tienda_id": None},
    "ingresos": {
        "ventas_brutas":      1500000.00,
        "menos_descuentos":     30000.00,
        "menos_devoluciones":   45000.00,
        "ingresos_netos":     1425000.00,
        "impuestos_cobrados":  120000.00,
        "num_ventas":               34,
        "num_devoluciones":          2,
    },
    "costo_ventas":       780000.00,
    "margen_bruto":       645000.00,
    "margen_bruto_pct":       45.26,
    "gastos_operativos": {
        "total": 200000.00,
        "detalle": [
            {"categoria": "arriendo",  "total": 120000.00, "cantidad": 1},
            {"categoria": "servicios", "total":  80000.00, "cantidad": 3},
        ],
    },
    "averias": {
        "perdidas_brutas":  50000.00,
        "valor_recuperado": 20000.00,
        "perdida_neta":     30000.00,
    },
    "utilidad_operativa":     415000.00,
    "utilidad_operativa_pct":     29.12,
}

# ── Construir el workbook ─────────────────────────────────────
try:
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Resultados"

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18

    r = 1
    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "ESTADO DE RESULTADOS", bold=True, bg=_NAVY, size=13)
    r += 1
    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, f"{data['periodo']['desde']} -> {data['periodo']['hasta']}",
         bold=False, bg=_NAVY, color="CCDDEE", size=10)
    r += 2

    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "INGRESOS", bold=True, bg=_BLUE, size=11)
    r += 1
    ing = data["ingresos"]
    rows_ing = [
        ("Ventas brutas",         ing["ventas_brutas"],      False),
        ("(-) Descuentos",       -ing["menos_descuentos"],   False),
        ("(-) Devoluciones",     -ing["menos_devoluciones"], False),
        (">> Ingresos netos",     ing["ingresos_netos"],     True),
        ("Impuestos cobrados",    ing["impuestos_cobrados"], False),
    ]
    for label, val, bold in rows_ing:
        _hdr(ws, r, 1, label, bold=bold, bg=_GRAY if bold else None,
             color="000000" if bold else "FFFFFF")
        _val(ws, r, 2, val, bold=bold,
             color=_GREEN if val >= 0 else _RED, bg=_GRAY if bold else None)
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "COSTO DE VENTAS", bold=True, bg=_BLUE, size=11)
    r += 1
    _hdr(ws, r, 1, "Costo de ventas (COGS)", color="000000")
    _val(ws, r, 2, -data["costo_ventas"], color=_RED)
    r += 1
    _hdr(ws, r, 1, ">> Margen bruto", bold=True, bg=_GRAY, color="000000")
    _val(ws, r, 2, data["margen_bruto"], bold=True,
         color=_GREEN if data["margen_bruto"] >= 0 else _RED, bg=_GRAY)
    r += 1
    _hdr(ws, r, 1, "Margen bruto %", bold=False, color="000000")
    _val(ws, r, 2, data["margen_bruto_pct"] / 100, fmt=_FMT_PCT, color="555555")
    r += 2

    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "GASTOS OPERATIVOS", bold=True, bg=_BLUE, size=11)
    r += 1
    for g in data["gastos_operativos"]["detalle"]:
        _hdr(ws, r, 1, f"  {g['categoria'].capitalize()}", color="000000")
        _val(ws, r, 2, -g["total"], color=_RED)
        r += 1
    _hdr(ws, r, 1, ">> Total gastos", bold=True, bg=_GRAY, color="000000")
    _val(ws, r, 2, -data["gastos_operativos"]["total"], bold=True, color=_RED, bg=_GRAY)
    r += 2

    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "RESULTADO", bold=True, bg=_NAVY, size=12)
    r += 1
    uo = data["utilidad_operativa"]
    _hdr(ws, r, 1, "UTILIDAD OPERATIVA", bold=True, bg=_NAVY, size=12)
    _val(ws, r, 2, uo, bold=True,
         color=_GREEN if uo >= 0 else _RED, bg=_NAVY)
    r += 1
    _hdr(ws, r, 1, "Utilidad %", color="000000")
    _val(ws, r, 2, data["utilidad_operativa_pct"] / 100, fmt=_FMT_PCT, color="555555")

    salida = "test_reporte.xlsx"
    wb.save(salida)
    size = os.path.getsize(salida)
    print(f"\n✅  Archivo generado: {salida}  ({size} bytes)")
    print("   Abrelo directamente desde el explorador de Windows para verificar.")

except Exception as e:
    import traceback
    print(f"\n❌  ERROR durante la generacion:")
    print(traceback.format_exc())
    sys.exit(1)
