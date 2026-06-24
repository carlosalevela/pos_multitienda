"""
Script de diagnóstico: genera el Excel directamente (sin Django)
para verificar si openpyxl produce un archivo válido.
Corre con: python test_excel.py
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_NAVY  = "1E3A5F"
_BLUE  = "2E6DA4"
_GRAY  = "F2F2F2"
_GREEN = "1E7E34"
_RED   = "B22222"
_FMT_CUR = '#,##0.00'
_FMT_PCT = '0.00%'


def _hdr(ws, row, col, value, bold=True, bg=None, color="FFFFFF", size=11, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color, size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell


def _val(ws, row, col, value, bold=False, color="000000", fmt=_FMT_CUR, bg=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


def _build_estado_resultados_sheet(ws, data):
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18

    r = 1
    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "ESTADO DE RESULTADOS", bold=True, bg=_NAVY, size=13)
    r += 1
    ws.merge_cells(f"A{r}:B{r}")
    periodo = f"{data['periodo']['desde']}  →  {data['periodo']['hasta']}"
    _hdr(ws, r, 1, periodo, bold=False, bg=_NAVY, color="CCDDEE", size=10)
    r += 2

    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "INGRESOS", bold=True, bg=_BLUE, size=11)
    r += 1
    ing = data["ingresos"]
    rows_ing = [
        ("Ventas brutas",         ing["ventas_brutas"],      False),
        ("(-) Descuentos",       -ing["menos_descuentos"],   False),
        ("(-) Devoluciones",     -ing["menos_devoluciones"], False),
        (">> Ingresos netos",     ing["ingresos_netos"],     True),
        ("Impuestos cobrados",    ing["impuestos_cobrados"], False),
    ]
    for label, val, bold in rows_ing:
        _hdr(ws, r, 1, label, bold=bold, bg=_GRAY if bold else None)
        color = _GREEN if val >= 0 else _RED
        _val(ws, r, 2, val, bold=bold, color=color, bg=_GRAY if bold else None)
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "COSTO DE VENTAS", bold=True, bg=_BLUE, size=11)
    r += 1
    _hdr(ws, r, 1, "Costo de ventas (COGS)")
    _val(ws, r, 2, -data["costo_ventas"], color=_RED)
    r += 1
    _hdr(ws, r, 1, ">> Margen bruto", bold=True, bg=_GRAY)
    _val(ws, r, 2, data["margen_bruto"], bold=True,
         color=_GREEN if data["margen_bruto"] >= 0 else _RED, bg=_GRAY)
    r += 1
    _hdr(ws, r, 1, "  Margen bruto %", bold=False, color="555555")
    _val(ws, r, 2, data["margen_bruto_pct"] / 100, fmt=_FMT_PCT, color="555555")
    r += 2

    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "GASTOS OPERATIVOS", bold=True, bg=_BLUE, size=11)
    r += 1
    for g in data["gastos_operativos"]["detalle"]:
        cat = g["categoria"] or "Sin categoría"
        _hdr(ws, r, 1, f"  {cat.capitalize()}")
        _val(ws, r, 2, -g["total"], color=_RED)
        r += 1
    _hdr(ws, r, 1, ">> Total gastos operativos", bold=True, bg=_GRAY)
    _val(ws, r, 2, -data["gastos_operativos"]["total"], bold=True, color=_RED, bg=_GRAY)
    r += 2

    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "AVERÍAS / DAÑOS", bold=True, bg=_BLUE, size=11)
    r += 1
    av = data["averias"]
    _hdr(ws, r, 1, "  Pérdidas brutas por daños")
    _val(ws, r, 2, -av["perdidas_brutas"], color=_RED)
    r += 1
    _hdr(ws, r, 1, "  (+) Valor recuperado")
    _val(ws, r, 2, av["valor_recuperado"], color=_GREEN)
    r += 1
    _hdr(ws, r, 1, ">> Pérdida neta por averías", bold=True, bg=_GRAY)
    _val(ws, r, 2, -av["perdida_neta"], bold=True, color=_RED, bg=_GRAY)
    r += 2

    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws, r, 1, "RESULTADO", bold=True, bg=_NAVY, size=11)
    r += 1
    uo = data["utilidad_operativa"]
    _hdr(ws, r, 1, "UTILIDAD OPERATIVA", bold=True, bg=_NAVY, color="FFFFFF", size=12)
    _val(ws, r, 2, uo, bold=True,
         color=_GREEN if uo >= 0 else _RED, bg=_NAVY, fmt=_FMT_CUR)
    ws.cell(row=r, column=2).font = Font(bold=True, color=_GREEN if uo >= 0 else _RED, size=12)
    r += 1
    _hdr(ws, r, 1, "  Margen operativo %", bold=False, color="AAAAAA")
    _val(ws, r, 2, data["utilidad_operativa_pct"] / 100, fmt=_FMT_PCT, color="AAAAAA")
    r += 1
    _hdr(ws, r, 1,
         f"  N° ventas: {data['ingresos']['num_ventas']}  |  "
         f"N° devoluciones: {data['ingresos']['num_devoluciones']}")


# ── Datos de prueba ───────────────────────────────────────────────
test_data = {
    "periodo": {"desde": "2024-01-01", "hasta": "2024-01-31"},
    "ingresos": {
        "ventas_brutas":      5000.0,
        "menos_descuentos":    100.0,
        "menos_devoluciones":   50.0,
        "ingresos_netos":     4850.0,
        "impuestos_cobrados":  200.0,
        "num_ventas":           25,
        "num_devoluciones":      2,
    },
    "costo_ventas":    2000.0,
    "margen_bruto":    2850.0,
    "margen_bruto_pct":  58.76,
    "gastos_operativos": {
        "total": 1000.0,
        "detalle": [
            {"categoria": "nomina",   "total": 600.0, "cantidad": 3},
            {"categoria": "arriendo", "total": 400.0, "cantidad": 1},
        ],
    },
    "averias": {
        "perdidas_brutas":  100.0,
        "valor_recuperado":  20.0,
        "perdida_neta":      80.0,
    },
    "utilidad_operativa":     1770.0,
    "utilidad_operativa_pct":   36.49,
}

if __name__ == "__main__":
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Resultados"

    try:
        _build_estado_resultados_sheet(ws, test_data)
        print("✅ _build_estado_resultados_sheet: OK")
    except Exception as e:
        print(f"❌ Error en _build_estado_resultados_sheet: {e}")
        raise

    buf = io.BytesIO()
    try:
        wb.save(buf)
        print("✅ wb.save: OK")
    except Exception as e:
        print(f"❌ Error en wb.save: {e}")
        raise

    data = buf.getvalue()
    print(f"   Tamaño: {len(data)} bytes")
    print(f"   Magic bytes: {data[:4].hex()} ({'OK (PK)' if data[:2] == b'PK' else 'INVALIDO'})")

    out_path = "test_estado_resultados.xlsx"
    with open(out_path, "wb") as f:
        f.write(data)
    print(f"   Guardado en: {out_path}")
    print()
    print("Abre test_estado_resultados.xlsx en Excel para verificar.")
    print("Si abre bien → el problema está en la respuesta Django, no en el Excel.")
    print("Si no abre → hay un bug en la generación del Excel.")
